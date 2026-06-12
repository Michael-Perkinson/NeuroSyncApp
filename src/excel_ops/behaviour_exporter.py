"""
Pure Excel export functions for photometry-behaviour alignment.

No UI framework dependencies — all Tkinter widget values must be resolved
by the caller before passing into these functions.
"""

from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import pandas as pd

from src.processing.behaviour_parser import retrieve_static_params, truncate_sheet_title


# ---------------------------------------------------------------------------
# DataFrame building
# ---------------------------------------------------------------------------

def _combine_signal_parts(start_data, end_data) -> pd.Series:
    return pd.concat(
        [start_data.reset_index(drop=True), end_data.reset_index(drop=True)],
        axis=0,
    ).reset_index(drop=True)


def _build_df_from_instance_time_ranges(
    behaviour_name: str,
    data_list: list[tuple],
    instance_time_ranges: list[pd.Series],
) -> pd.DataFrame:
    """Build a behaviour frame aligned by actual relative timestamps."""
    df_combined = pd.DataFrame()

    for instance_index, ((start_data, end_data), time_range) in enumerate(
        zip(data_list, instance_time_ranges), start=1
    ):
        combined_data = _combine_signal_parts(start_data, end_data)
        time_values = pd.Series(time_range).reset_index(drop=True)
        row_count = min(len(combined_data), len(time_values))
        column_name = f"{behaviour_name}_{instance_index}"
        frame = pd.DataFrame(
            {
                "Time (s)": time_values.iloc[:row_count].reset_index(drop=True),
                column_name: combined_data.iloc[:row_count].reset_index(drop=True),
            }
        ).dropna(subset=["Time (s)"])
        frame = frame.groupby("Time (s)", as_index=False).mean(numeric_only=True)

        if df_combined.empty:
            df_combined = frame
        else:
            df_combined = pd.merge(df_combined, frame, on="Time (s)", how="outer")

    if df_combined.empty:
        return pd.DataFrame({"Time (s)": []})

    df_combined = df_combined.sort_values("Time (s)").reset_index(drop=True)
    instance_columns = [col for col in df_combined.columns if col != "Time (s)"]
    if len(instance_columns) > 1:
        df_combined["Mean"] = df_combined[instance_columns].mean(axis=1)
        df_combined["SEM"] = df_combined[instance_columns + ["Mean"]].sem(axis=1)

    return df_combined


def create_df_for_behaviours(
    behaviours_results: dict,
    sorted_behaviours: list[str],
    time_ranges: dict,
    combine_csv: bool,
    selected_column: str,
    checkbox_state: bool,
    folder_path: str,
) -> list[tuple[pd.DataFrame, str]]:
    """Build per-behaviour DataFrames and optionally write individual CSVs.

    Parameters
    ----------
    behaviours_results:
        ``{behaviour_name: [(start_data, end_data), ...]}``.
    sorted_behaviours:
        Behaviour names in display order.
    time_ranges:
        ``{behaviour_name: pd.Series}`` — time axis for each behaviour.
    combine_csv:
        If ``True``, collect DataFrames into the returned list instead of
        writing individual CSV files.
    selected_column:
        Signal column name (used only when *combine_csv* is ``False``).
    checkbox_state:
        Whether z-score / baseline mode is active (affects CSV filename).
    folder_path:
        Output directory for individual CSV files.

    Returns
    -------
    list of ``(DataFrame, sheet_name)`` tuples — populated only when
    *combine_csv* is ``True``.
    """
    df_list: list[tuple[pd.DataFrame, str]] = []

    for behaviour_name in sorted_behaviours:
        data_list = behaviours_results[behaviour_name]
        behaviour_time_ranges = time_ranges[behaviour_name]

        if (
            isinstance(behaviour_time_ranges, list)
            and len(behaviour_time_ranges) >= len(data_list)
        ):
            df_combined = _build_df_from_instance_time_ranges(
                behaviour_name, data_list, behaviour_time_ranges
            )
        else:
            df_combined = pd.DataFrame()

            for instance_index, (start_data, end_data) in enumerate(data_list, start=1):
                combined_data = _combine_signal_parts(start_data, end_data)
                df_combined[f"{behaviour_name}_{instance_index}"] = combined_data

            if len(df_combined.columns) > 1:
                df_combined["Mean"] = df_combined.mean(axis=1)
                df_combined["SEM"] = df_combined.sem(axis=1)

            time_range = pd.Series(behaviour_time_ranges)
            df_combined.insert(
                0, "Time (s)", time_range[: len(df_combined)].reset_index(drop=True)
            )

            for col in df_combined.columns:
                df_combined[col] = (
                    df_combined[col].reindex(time_range.index).reset_index(drop=True)
                )

        sheet_name = truncate_sheet_title(behaviour_name)
        sheet_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)

        if combine_csv:
            df_list.append((df_combined, sheet_name))
        else:
            results_file_name = f"{sheet_name}_{selected_column}_raw.csv"
            if checkbox_state:
                results_file_name = results_file_name.replace("_raw.csv", "_baseline_raw.csv")
            (Path(folder_path) / results_file_name).parent.mkdir(parents=True, exist_ok=True)
            df_combined.to_csv(Path(folder_path) / results_file_name, index=False)

    return df_list


# ---------------------------------------------------------------------------
# Binning
# ---------------------------------------------------------------------------

def process_and_bin_data(
    data_list: list[tuple],
    pre_behaviour_time: int,
    post_behaviour_time: int,
    bin_size: int,
    time_ranges: list[pd.Series] | None = None,
) -> tuple[list[str], int, list]:
    """Bin concatenated signal data for all instances of one behaviour.

    Parameters
    ----------
    data_list:
        ``[(start_data, end_data), ...]`` — one tuple per event instance.
    pre_behaviour_time, post_behaviour_time, bin_size:
        Integers (seconds).
    time_ranges:
        Optional actual relative time axis for each event instance. When
        present, samples are assigned to bins by timestamp rather than by row
        count, preserving scheduled-recording gaps.

    Returns
    -------
    bin_labels : list[str]
        Human-readable ``"start - end"`` strings for each bin.
    num_bins_total : int
        Total number of bins across the window.
    behaviour_instances_data : list
        Flat list of 1-D arrays (one per bin per instance).
    """
    behaviour_instances_data = []
    total_window = pre_behaviour_time + post_behaviour_time
    num_bins_total = total_window // bin_size
    bin_ranges = np.linspace(
        -pre_behaviour_time, post_behaviour_time, num_bins_total, endpoint=False
    )
    bin_labels = [
        f"{start} - {start + bin_size}"
        for start in sorted(bin_ranges, key=float)
    ]

    for instance_index, (start_data, end_data) in enumerate(data_list):
        combined = np.concatenate([start_data.values, end_data.values]).ravel()
        if time_ranges is not None and instance_index < len(time_ranges):
            instance_times = pd.Series(time_ranges[instance_index]).reset_index(
                drop=True
            )
            instance_times = instance_times.iloc[: len(combined)].to_numpy(dtype=float)
            for bin_start in bin_ranges:
                bin_end = bin_start + bin_size
                mask = (instance_times >= bin_start) & (instance_times < bin_end)
                behaviour_instances_data.append(combined[mask])
        else:
            for values in np.array_split(combined, num_bins_total):
                behaviour_instances_data.append(values)

    return bin_labels, num_bins_total, behaviour_instances_data


# ---------------------------------------------------------------------------
# Summary data
# ---------------------------------------------------------------------------

def generate_summary_data(
    sorted_behaviours: list[str],
    behaviours_results: dict,
    params: dict,
    metric_functions: dict,
    time_ranges: dict | None = None,
) -> pd.DataFrame:
    """Build the summary DataFrame with bin labels and per-metric rows.

    Parameters
    ----------
    sorted_behaviours:
        Behaviour names in display order.
    behaviours_results:
        ``{behaviour_name: [(start_data, end_data), ...]}``.
    params:
        Extraction params dict (keyed by behaviour name).
    metric_functions:
        ``{metric_name: callable}`` — caller resolves which metrics are
        enabled before passing this in.

    Returns
    -------
    pd.DataFrame
        Un-headered summary table ready for Excel export.
    """
    from src.processing.behavior_metrics import calculate_metrics_for_bins  # avoid circular

    summary_rows: list[list] = []

    for behaviour_name in sorted_behaviours:
        data_list = behaviours_results[behaviour_name]
        pre, post, bin_ = retrieve_static_params(params, behaviour_name)

        instance_time_ranges = time_ranges.get(behaviour_name) if time_ranges else None
        bin_labels, expected_num_bins, behaviour_instances_data = process_and_bin_data(
            data_list, pre, post, bin_, instance_time_ranges
        )

        summary_rows.append([behaviour_name] + bin_labels)

        instances_data = [
            behaviour_instances_data[i : i + expected_num_bins]
            for i in range(0, len(behaviour_instances_data), expected_num_bins)
        ]

        for metric_name, metric_func in metric_functions.items():
            metric_values = calculate_metrics_for_bins(
                instances_data, expected_num_bins, metric_name, metric_func
            )
            summary_rows.append([metric_name] + metric_values)

        summary_rows.append([""] * (1 + len(bin_labels)))

    return pd.DataFrame(summary_rows)


def prepare_combined_data(df_summary: pd.DataFrame) -> list[pd.DataFrame]:
    """Return ``[df_summary, separator_row]`` ready to concatenate."""
    separator = pd.DataFrame(np.nan, columns=df_summary.columns, index=[0])
    return [df_summary, separator]


# ---------------------------------------------------------------------------
# File-name helpers
# ---------------------------------------------------------------------------

def build_output_file_name(
    file_path: str,
    folder_path: str,
    selected_column: str,
    checkbox_state: bool,
    baseline_start: str = "",
) -> str:
    """Return the full path for the combined Excel output file.

    Parameters
    ----------
    file_path:
        Original photometry CSV path (stem used as base name).
    folder_path:
        Output directory.
    selected_column:
        Signal column name appended to the stem.
    checkbox_state:
        If ``True``, ``_baseline_<baseline_start>`` is appended.
    baseline_start:
        Raw string from the baseline-start entry widget (used verbatim).
    """
    stem = Path(file_path).stem
    name = f"{stem}_{selected_column}"
    if checkbox_state:
        name += f"_baseline_{baseline_start}"
    return str(Path(folder_path) / f"{name}.xlsx")


def build_separate_summary_file_name(
    file_path: str,
    folder_path: str,
    selected_column: str,
    checkbox_state: bool,
    baseline_start: str = "",
    use_binned_data: bool = False,
) -> str:
    """Return the full path for a non-combined summary CSV."""
    stem = Path(file_path).stem
    name = f"{stem}_{selected_column}"
    if checkbox_state:
        name += f"_baseline_{baseline_start}"
    if use_binned_data:
        name += "_binned"
    return str(Path(folder_path) / f"{name}_summary.csv")


# ---------------------------------------------------------------------------
# Non-binned metrics DataFrame
# ---------------------------------------------------------------------------

def build_non_binned_metrics_df(
    behaviours_results: dict,
    params: dict,
    metric_flags: dict,
    metric_functions: dict,
    checkbox_state: bool,
    baseline_start_time: float = 0.0,
    baseline_end_time: float = 0.0,
) -> pd.DataFrame:
    """Build the summary DataFrame for the non-binned export path.

    Parameters
    ----------
    behaviours_results:
        ``{behaviour_name: [(start_data, end_data), ...]}``.
    params:
        Extraction-params dict (keyed by behaviour name).
    metric_flags:
        Resolved bool flags: ``use_auc``, ``use_max_amp``, ``use_mean_dff``.
    metric_functions:
        ``{metric_name: callable}`` — same keys as metric_flags minus
        the ``use_`` prefix.
    checkbox_state:
        If ``True``, baseline columns are appended to the output.
    baseline_start_time, baseline_end_time:
        Values written into the baseline columns (used only when
        *checkbox_state* is ``True``).

    Returns
    -------
    pd.DataFrame
        Wide summary table ready for Excel / CSV export.
    """
    behaviours_metrics: dict = {}
    bin_size = params.get("bin_size")

    for behaviour_name in behaviours_results:
        for behaviour_instance in params[behaviour_name]:
            pre  = float(behaviour_instance["pre_behaviour_time"])
            post = float(behaviour_instance["post_behaviour_time"])

            for start_data, end_data in behaviours_results[behaviour_name]:
                data_dict: dict = {
                    "behaviour_name":    behaviour_name,
                    "pre_behaviour_time":  pre,
                    "post_behaviour_time": post,
                    "bin_size":            bin_size,
                }
                if metric_flags.get("use_auc"):
                    data_dict["start_auc"] = metric_functions["auc"](start_data)
                    data_dict["end_auc"]   = metric_functions["auc"](end_data)
                if metric_flags.get("use_max_amp"):
                    data_dict["start_max_amp"] = metric_functions["max_amp"](start_data)
                    data_dict["end_max_amp"]   = metric_functions["max_amp"](end_data)
                if metric_flags.get("use_mean_dff"):
                    data_dict["start_mean_dff"] = metric_functions["mean_dff"](start_data)
                    data_dict["end_mean_dff"]   = metric_functions["mean_dff"](end_data)

                bm = behaviours_metrics.setdefault(behaviour_name, {})
                scalar_keys = {"behaviour_name", "pre_behaviour_time",
                               "post_behaviour_time", "bin_size"}
                for key, val in data_dict.items():
                    if key not in scalar_keys:
                        bm.setdefault(key, []).append(val)
                for key in ("pre_behaviour_time", "post_behaviour_time", "bin_size"):
                    bm[key] = data_dict[key]

    def _add_empty(df: pd.DataFrame) -> pd.DataFrame:
        df[""] = ""
        return df

    dfs = []
    df_times = pd.DataFrame(
        [(k, v["pre_behaviour_time"], v["post_behaviour_time"])
         for k, v in behaviours_metrics.items()],
        columns=["Behaviour", "Pre Behaviour Time", "Post Behaviour Time"],
    )
    dfs.append(_add_empty(df_times))

    first = behaviours_metrics[next(iter(behaviours_metrics))]
    if "start_auc" in first:
        df_auc = pd.DataFrame(
            [(k, np.mean(v.get("start_auc", [np.nan])),
                 np.mean(v.get("end_auc",   [np.nan])))
             for k, v in behaviours_metrics.items()],
            columns=["Behaviour", "Start AUC", "End AUC"],
        )
        dfs.append(_add_empty(df_auc))
    if "start_max_amp" in first:
        df_max_amp = pd.DataFrame(
            [(k, np.mean(v.get("start_max_amp", [np.nan])),
                 np.mean(v.get("end_max_amp",   [np.nan])))
             for k, v in behaviours_metrics.items()],
            columns=["Behaviour", "Start Max AMP", "End Max AMP"],
        )
        dfs.append(_add_empty(df_max_amp))
    if "start_mean_dff" in first:
        df_mean_dff = pd.DataFrame(
            [(k, np.mean(v.get("start_mean_dff", [np.nan])),
                 np.mean(v.get("end_mean_dff",   [np.nan])))
             for k, v in behaviours_metrics.items()],
            columns=["Behaviour", "Start Mean dF/F", "End Mean dF/F"],
        )
        dfs.append(_add_empty(df_mean_dff))

    df_results = pd.concat(dfs, axis=1)

    if checkbox_state:
        df_results["Baseline Start Time"] = ""
        df_results["Baseline End Time"]   = ""
        df_results.at[0, "Baseline Start Time"] = baseline_start_time
        df_results.at[0, "Baseline End Time"]   = baseline_end_time

    return df_results


def format_non_binned_excel(output_file_name: str) -> None:
    """Remove borders from the header row of every sheet in *output_file_name*."""
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side
    wb = load_workbook(output_file_name)
    no_border = Border(
        left=Side(style="none"),
        right=Side(style="none"),
        top=Side(style="none"),
        bottom=Side(style="none"),
    )
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.border = no_border
    wb.save(output_file_name)


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

def save_to_excel(df_list: list[tuple[pd.DataFrame, str]], output_file_name: str) -> None:
    """Write each ``(DataFrame, sheet_name)`` pair to *output_file_name*."""
    with pd.ExcelWriter(output_file_name, engine="openpyxl") as writer:
        for df, sheet_name in df_list:
            df.to_excel(writer, sheet_name=sheet_name.replace("/", "_"), index=False)


def export_combined_csv(
    df_summary: pd.DataFrame,
    df_list: list[tuple[pd.DataFrame, str]],
    output_file_name: str,
    behaviours_results: dict,
) -> None:
    """Concatenate summary + per-behaviour frames and write one Excel file.

    Parameters
    ----------
    df_summary:
        Summary DataFrame produced by :func:`generate_summary_data`.
    df_list:
        Per-behaviour DataFrames (already built by
        :func:`create_df_for_behaviours`).  Mutated in-place: the Event
        Duration frame and Summary Results frame are prepended.
    output_file_name:
        Full path for the output ``.xlsx`` file.
    behaviours_results:
        Used by :func:`format_excel` to bold behaviour-name rows.
    """
    combined_parts = prepare_combined_data(df_summary)
    df_summary_combined = pd.concat(combined_parts, ignore_index=True)
    df_summary_combined.columns = [""] * df_summary_combined.shape[1]
    df_summary_combined.columns.name = None

    df_list.insert(0, (df_summary_combined, "Summary Results"))

    save_to_excel(df_list, output_file_name)
    format_excel(output_file_name, behaviours_results)


def export_separate_csv(
    df_summary: pd.DataFrame,
    output_file_name: str,
) -> None:
    """Write *df_summary* as a headerless CSV to *output_file_name*."""
    df_summary.to_csv(output_file_name, index=False, header=False)


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

def bold_first_row(ws) -> None:
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)


def bold_first_column(ws) -> None:
    from openpyxl.styles import Font
    for cell in ws["A"]:
        cell.font = Font(bold=True)


def bold_behavior_name_rows(ws, behaviours_results: dict) -> None:
    from openpyxl.styles import Font
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
        if row[0].value in behaviours_results:
            for cell in row:
                cell.font = Font(bold=True)


def remove_borders(ws) -> None:
    from openpyxl.styles import Border, Side
    no_border = Border(
        left=Side(style="none"),
        right=Side(style="none"),
        top=Side(style="none"),
        bottom=Side(style="none"),
    )
    for cell in ws[1]:
        cell.border = no_border


def format_excel(output_file_name: str, behaviours_results: dict) -> None:
    """Apply bold/border formatting to an already-written Excel workbook."""
    from openpyxl import load_workbook
    wb = load_workbook(output_file_name)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == "Event Duration":
            bold_first_row(ws)
            bold_first_column(ws)
        else:
            if sheet_name == wb.sheetnames[0]:
                ws.delete_rows(1)
            bold_first_column(ws)
            bold_behavior_name_rows(ws, behaviours_results)
        remove_borders(ws)
    wb.save(output_file_name)
