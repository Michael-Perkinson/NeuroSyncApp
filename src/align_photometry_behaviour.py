import json
import os
import re
import tkinter as tk
import tkinter.font as tkf
import uuid
import copy
import math
import datetime
from tkinter import colorchooser, filedialog, messagebox, ttk
import matplotlib
import traceback
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.patches import Rectangle
from matplotlib.ticker import AutoLocator, MultipleLocator
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
from scipy.integrate import simpson

from behaviour_event_input_frame import BehaviourInputFrame
from data_selection_frame import DataSelectionFrame
from static_inputs_frame import StaticInputsFrame
from window_utils import center_window_on_screen
from app_settings_manager import AppSettingsManager
from export_options_container import ExportOptionsContainer
from graph_settings_container import GraphSettingsContainer


class DataProcessingSingleInstance(ttk.Frame):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        self.initialize_attributes(parent)
        self.create_widgets()
        self.setup_notebooks()
        self.create_tabs()
        self.initialize_graph_settings()
        self.create_graphs_container(self.graph_tab)
        self.create_table_container(self.table_tab)
        self.configure_grid_weights()

    def initialize_attributes(self, parent):
        """
        Initialize instance variables.

        Parameters:
        - parent: The parent widget.
        """
        self.init_notebooks()
        self.init_file_vars()
        self.init_time_vars()
        self.init_display_vars()
        self.init_behaviour_vars()

        if isinstance(parent, tk.Tk):
            parent.title("Data Processing")

        self.main_frame = ttk.Frame(self, style='Bordered.TFrame')
        self.main_frame.grid(row=0, column=0, columnspan=4,
                             padx=10, sticky=tk.NSEW)
        self.grid(row=0, column=0, sticky="nsew")

    def init_notebooks(self):
        """Initialize notebook attributes."""
        self.notebook_graphs = ttk.Notebook(self)

    def init_file_vars(self):
        """Initialize file-related variables."""
        self.file_path = ""
        self.file_path_var = tk.StringVar()
        self.is_file_parsed = False

    def init_time_vars(self):
        """Initialize time-related variables."""
        self.start_time = 0.0
        self.end_time = 0.0
        self.pre_behaviour_time_var = tk.StringVar()
        self.post_behaviour_time_var = tk.StringVar()
        self.synchronize_start_time_var = tk.StringVar()
        self.duration_box_placement_var = tk.StringVar(value='1')

    def init_display_vars(self):
        """Initialize display-related variables."""
        self.selected_column = ""
        self.selected_column_var = tk.StringVar()
        self.column_titles = []
        self.table_treeview = None
        self.figure_canvas = None
        self.figure_display_dropdown = None
        self.figure_display_choices = None
        self.checkbox_state = False
        self.warning_shown = False

    def init_behaviour_vars(self):
        """Initialize behaviour-related variables."""
        self.unique_behaviours = []
        self.behaviour_colors = {}
        self.behaviour_boxes = {}
        self.adjusted_behaviour_dataframes = {}
        self.behaviour_display_status = {}
        self.data_already_adjusted = False
        self.behaviours_results_auc_binned = []
        self.behaviours_results_max_amp_binned = []
        self.behaviours_results_mean_dff_binned = []
        self.default_column_names = {
            'Behaviours/events': '', 'Start Time': '', 'End Time': ''}
        self.display_duration_box_var = tk.BooleanVar(value=True)
        self.num_instances_box_var = tk.BooleanVar(value=True)
        self.time_unit_var = tk.StringVar(value="minutes")
        self.time_input_unit_var = tk.StringVar(value="seconds")
        self.x_gridlines_var = tk.StringVar()
        self.y_gridlines_var = tk.StringVar()
        self.current_table_key = None
        self.first_offset_time = None
        self.baseline_button_pressed = False
        self.mouse_name = None
        self.column_dropdown = None
        self.dataframe = None
        self.z_score_computed = False
        self.previous_baseline_start = None
        self.previous_baseline_end = None
        self.x_axis_min_var = tk.StringVar()
        self.x_axis_max_var = tk.StringVar()
        self.y_axis_min_var = tk.StringVar()
        self.y_axis_max_var = tk.StringVar()
        self.behaviour_coding_file_var = tk.StringVar()
        self.settings_manager = AppSettingsManager(
            app_type="align_photometry_and_behaviour_app")
        self.app_name = "align_photometry_and_behaviour_app"
        self.settings_manager.load_variables()
        self.xlim_max = None
        self.xlim_min = None
        self.tables = {}

    def create_widgets(self):
        self.create_data_selection_frame()
        self.create_behaviour_event_input_frame()
        self.create_static_inputs_frame()
        self.define_custom_styles()

    def create_data_selection_frame(self):
        """Create and configure the DataSelectionFrame."""
        self.data_selection_frame = DataSelectionFrame(
            self.main_frame,
            width=500,
            settings_manager=self.settings_manager,
            figure_display_callback=self.handle_figure_display_selection,
            new_data_file_callback=self.handle_new_data_file,
            figure_display_dropdown=self.figure_display_dropdown,
            figure_display_choices=self.figure_display_choices,
            update_table_from_frame_callback=self.update_table_from_frame,
        )
        self.data_selection_frame.grid(
            row=0, column=0, padx=10, pady=10, sticky=tk.NSEW)

    def create_behaviour_event_input_frame(self):
        """Create and configure the BehaviourInputFrame."""
        self.behaviour_event_input_frame = BehaviourInputFrame(
            self.main_frame,
            width=450,
            select_column_names_callback=self.select_column_names,
            select_event_file_callback=self.select_event_file,
        )
        self.behaviour_event_input_frame.grid(
            row=0, column=2, padx=10, pady=10, sticky=tk.NSEW)
        self.behaviour_event_input_frame.columnconfigure(0, weight=0)
        self.behaviour_event_input_frame.columnconfigure(1, weight=1)

    def create_static_inputs_frame(self):
        """Create and configure the StaticInputsFrame."""
        self.static_inputs_frame = StaticInputsFrame(self.main_frame,
                                                     width=230,
                                                     save_inputs_callback=self.save_inputs)
        self.static_inputs_frame.grid(
            row=0, column=3, padx=10, pady=10, sticky=tk.NSEW)

    def define_custom_styles(self):
        """Define custom styles for ttk widgets."""
        style = ttk.Style()
        style.configure('Custom.TFrame', background='snow')
        style.configure('Bordered.TFrame', background='snow',
                        borderwidth=2, relief='solid', bordercolor='black')
        style.configure('Custom.TCheckbutton', background='snow', foreground='black', font=('Helvetica', 10), compound='left', padding=(20, 0, 0, 0),
                        wraplength=150)
        style.configure("CustomScale.TScale", background="snow",
                        troughcolor="lightgray", gripcount=0)
        style.configure("NoBorder.TFrame", background="snow",
                        borderwidth=0, padding=0)
        style.configure("CustomNotebook.TNotebook", background="snow")
        style.configure("CustomNotebook.TFrame", background="snow")
        style.configure("Custom.TMenubutton", background='snow', width=10, font=('Helvetica', 8), bd=1, relief="solid",
                        indicatoron=True)

    def setup_notebooks(self):
        """Setup the main and settings notebooks."""
        self.notebook_graphs = ttk.Notebook(
            self, style="CustomNotebook.TNotebook")
        self.notebook_graphs.grid(
            row=1, column=0, columnspan=3, padx=10, sticky=tk.NSEW)

        self.notebook_settings = ttk.Notebook(
            self, style="CustomNotebook.TNotebook", height=520, width=300)
        self.notebook_settings.grid(row=1, column=3, padx=10, sticky=tk.NSEW)

    def create_tabs(self):
        """Create and add tabs to the notebooks."""
        self.graph_settings_tab = ttk.Frame(self.notebook_settings)
        self.export_options_tab = ttk.Frame(self.notebook_settings)

        self.notebook_settings.add(
            self.graph_settings_tab, text="Graph Settings")
        self.notebook_settings.add(
            self.export_options_tab, text="Export Options")

    def initialize_graph_settings(self):
        """Initialize graph settings and export options."""
        self.graph_settings_container_instance = GraphSettingsContainer(
            self.graph_settings_tab,
            widgets_to_include=[
                'line_width', 'box_height', 'alpha',
                'bar_graph_size', 'onset_line_thickness', 'onset_line_style',
                'color_buttons', 'checkboxes', 'axis_range',
                'zero_x_axis_to_behaviour', 'graph_time_labels'
            ],
            app_name=self.app_name,
            settings_manager=self.settings_manager,
            refresh_graph_display_callback=self.refresh_graph_display,
            update_duration_box_callback=self.update_duration_box,
            handle_behaviour_change_callback=self.handle_behaviour_change,
            load_variables_callback=self.settings_manager.load_variables,
            save_variables_callback=self.settings_manager.save_variables,
            create_behaviour_options_callback=self.create_behaviour_options,
            update_box_colors_callback=self.update_box_colors_and_behaviour_options,
            save_and_close_axis_callback=self.save_and_close,
        )

        self.export_options_container = ExportOptionsContainer(
            self.export_options_tab,
            file_path_var=self.file_path_var,
            settings_manager=self.settings_manager,
            extract_button_click_handler=self.extract_button_click_handler,
            save_image=self.save_image
        )

        self.settings_manager.update_graph_settings_container(
            self.graph_settings_container_instance)
        self.settings_manager.update_export_options_container(
            self.export_options_container)
        self.graph_settings_container_instance.complete_initialization()

        self.graph_settings_tab.grid_rowconfigure(0, weight=1)
        self.graph_settings_tab.grid_columnconfigure(0, weight=1)
        self.export_options_tab.grid_rowconfigure(0, weight=1)
        self.export_options_tab.grid_columnconfigure(0, weight=1)

        # Create tabs
        self.graph_tab = ttk.Frame(
            self.notebook_graphs, style="CustomNotebook.TFrame")
        self.table_tab = ttk.Frame(
            self.notebook_graphs, style="CustomNotebook.TFrame")

        self.notebook_graphs.add(self.graph_tab, text="Graph")
        self.notebook_graphs.add(self.table_tab, text="Table")

    def configure_grid_weights(self):
        """Configure row and column weights for the frames."""
        self.graph_tab.grid_rowconfigure(0, weight=1)
        self.graph_tab.grid_columnconfigure(0, weight=1)
        self.table_tab.grid_rowconfigure(0, weight=1)
        self.table_tab.grid_columnconfigure(0, weight=1)

    def create_widgets(self):
        """Create and initialize widgets."""
        self.data_selection_frame = DataSelectionFrame(
            self.main_frame,
            width=500,
            settings_manager=self.settings_manager,
            figure_display_callback=self.handle_figure_display_selection,
            new_data_file_callback=self.handle_new_data_file,
            figure_display_dropdown=self.figure_display_dropdown,
            figure_display_choices=self.figure_display_choices,
            update_table_from_frame_callback=self.update_table_from_frame,
        )
        self.data_selection_frame.grid(
            row=0, column=0, padx=10, pady=10, sticky=tk.NSEW)

        self.behaviour_event_input_frame = BehaviourInputFrame(
            self.main_frame,
            width=450,
            select_column_names_callback=self.select_column_names,
            select_event_file_callback=self.select_event_file,
        )
        self.behaviour_event_input_frame.grid(
            row=0, column=2, padx=10, pady=10, sticky=tk.NSEW)

        self.behaviour_event_input_frame.columnconfigure(0, weight=0)
        self.behaviour_event_input_frame.columnconfigure(1, weight=1)

        self.static_inputs_frame = StaticInputsFrame(self.main_frame,
                                                     width=230,
                                                     save_inputs_callback=self.save_inputs)
        self.static_inputs_frame.grid(
            row=0, column=3, padx=10, pady=10, sticky=tk.NSEW)

        # Define custom styles for ttk widgets
        style = ttk.Style()
        style.configure('Custom.TFrame', background='snow')
        style.configure('Bordered.TFrame', background='snow',
                        borderwidth=2, relief='solid', bordercolor='black')
        style.configure('Custom.TCheckbutton', background='snow', foreground='black', font=('Helvetica', 10), compound='left', padding=(20, 0, 0, 0),
                        wraplength=150)
        style.configure("CustomScale.TScale", background="snow",
                        troughcolor="lightgray", gripcount=0)
        style.configure("NoBorder.TFrame", background="snow",
                        borderwidth=0, padding=0)
        style.configure("CustomNotebook.TNotebook", background="snow")
        style.configure("CustomNotebook.TFrame", background="snow")
        style.configure("Custom.TMenubutton", background='snow', width=10, font=('Helvetica', 8), bd=1, relief="solid",
                        indicatoron=True)

    def create_graphs_container(self, frame):
        """
        Create and initialize the graphs container.

        Parameters:
        - frame: The parent frame to contain the graphs.
        """
        # Graphs container frame
        graphs_container_frame = ttk.Frame(
            frame, style='NoBorder.TFrame', borderwidth=2, relief='solid')
        graphs_container_frame.grid(
            row=0, column=0, columnspan=3, padx=10, pady=10, sticky=tk.NSEW)
        graphs_container_frame.columnconfigure(0, weight=1)
        graphs_container_frame.rowconfigure(0, weight=1)

        color_button = tk.Button(graphs_container_frame, text="Main Trace Colour",
                                 bg='lightblue', command=self.pick_trace_color)
        color_button.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)

        sem_color_button = tk.Button(graphs_container_frame, text="SEM Colour", bg='lightblue',
                                     command=self.pick_sem_color)
        sem_color_button.grid(row=0, column=1, padx=10, pady=10, sticky=tk.W)

        self.selected_behaviour = tk.StringVar(
            value="Choose behaviour to plot")

        def behaviour_selection_changed(*args):
            """Callback function for when the behaviour selection dropdown is changed."""

            # Check if self.selected_behaviour is an empty string
            if self.selected_behaviour.get() == "":
                return  # Exit the function if it is empty

            if self.selected_column_var.get() == "Behaviour Mean and SEM":
                self.handle_figure_display_selection(None)
            else:
                self.figure_display_dropdown.set("Behaviour Mean and SEM")
                self.handle_figure_display_selection(None)

        self.selected_behaviour.trace("w", behaviour_selection_changed)

        self.behaviour_choice_graph = ttk.Combobox(
            graphs_container_frame, state="readonly", width=30, textvariable=self.selected_behaviour)
        self.behaviour_choice_graph.grid(row=0, column=2, padx=5, sticky=tk.W)
        self.behaviour_choice_graph.configure(
            state=tk.DISABLED)  # Disable the dropdown initially

        self.figure_display_choices = [
            "Full Trace Display", "Single Row Display", "Behaviour Mean and SEM"]
        self.figure_display_dropdown = ttk.Combobox(
            graphs_container_frame, state="readonly", values=self.figure_display_choices, width=30)
        self.figure_display_dropdown.grid(row=0, column=3, padx=5, sticky=tk.W)
        self.figure_display_dropdown.bind(
            "<<ComboboxSelected>>", self.handle_figure_display_selection)
        # Set the initial value to "Full Trace Display"
        self.figure_display_dropdown.set(self.figure_display_choices[0])
        self.data_selection_frame.set_figure_display_dropdown(
            self.figure_display_dropdown)
        self.data_selection_frame.set_figure_display_choices(
            self.figure_display_choices)

        # Create a canvas to display the graphs
        self.graph_canvas = tk.Canvas(
            graphs_container_frame, bg='snow', highlightthickness=1)
        self.graph_canvas.grid(row=1, column=0, columnspan=4, sticky=tk.NSEW)

        graphs_container_frame.grid_columnconfigure(0, weight=1)
        graphs_container_frame.grid_rowconfigure(0, weight=0)
        graphs_container_frame.grid_columnconfigure(0, weight=1)
        graphs_container_frame.grid_rowconfigure(1, weight=1)

    def create_table_container(self, frame):
        """
        Create and initialize the table container within the specified frame.

        This method sets up a Treeview inside a scrollable canvas, allowing for horizontal and vertical scrolling.
        It also configures the column headers, widths, and binds events for row selection and column sorting.

        Parameters:
        frame (ttk.Frame): The parent frame where the table container will be placed.
        """
        # Create the table container frame
        table_container_frame = ttk.Frame(
            frame, style='NoBorder.TFrame')  # Set a minimum or fixed width
        table_container_frame.grid(
            row=0, column=0, columnspan=3, padx=10, pady=10, sticky=tk.NSEW)

        # Create a horizontal scrollbar for the table, linked to the canvas
        table_hscrollbar = ttk.Scrollbar(
            table_container_frame, orient=tk.HORIZONTAL)
        # Pack the scrollbar first so it stays at the bottom
        table_hscrollbar.pack(side=tk.BOTTOM, fill=tk.X)

        # Add a vertical scrollbar for the canvas
        self.table_vscrollbar = ttk.Scrollbar(
            table_container_frame, orient="vertical")
        self.table_vscrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.table_canvas = tk.Canvas(table_container_frame, xscrollcommand=table_hscrollbar.set,
                                      yscrollcommand=self.table_vscrollbar.set, height=430)
        self.table_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Create a frame inside the canvas to hold the Treeview
        table_scroll_frame = ttk.Frame(self.table_canvas)
        self.table_canvas.create_window(
            (0, 0), window=table_scroll_frame, anchor='nw')

        # Bind a function to modify the scroll region
        def configure_scroll_region(event):
            """Configure the scroll region to encompass the full height of the frame."""
            self.table_canvas.configure(
                scrollregion=self.table_canvas.bbox("all"), height=420)

        table_scroll_frame.bind("<Configure>", configure_scroll_region)

        # Create the Treeview inside the scrollable frame
        self.table_treeview = ttk.Treeview(table_scroll_frame,
                                           columns=['file_name', 'column_title', 'behaviour_name', 'behaviour_type',
                                                    'pre_behaviour_time', 'post_behaviour_time', 'bin_size',
                                                    'start_time', 'end_time'],
                                           show='headings', name='treeview')

        # Set the height of the table to control the number of visible rows
        table_height = 30  # Set the initial height to 30 rows
        self.table_treeview.configure(height=table_height)

        # Pack the Treeview to make it visible and fill the available space
        self.table_treeview.pack(fill=tk.BOTH, expand=True)

        # Define table columns without lambda functions
        self.table_treeview.heading('file_name', text='File Name')
        self.table_treeview.heading('column_title', text='Column Title')
        self.table_treeview.heading('behaviour_name', text='Behaviour Name')
        self.table_treeview.heading('behaviour_type', text='Behaviour Type')
        self.table_treeview.heading(
            'pre_behaviour_time', text='Pre-behaviour Time')
        self.table_treeview.heading(
            'post_behaviour_time', text='Post-behaviour Time')
        self.table_treeview.heading('bin_size', text='Bin Size')
        self.table_treeview.heading('start_time', text='Start Time')
        self.table_treeview.heading('end_time', text='End Time')

        # Adjust column widths based on content
        columns = ['file_name', 'column_title', 'behaviour_name', 'behaviour_type', 'pre_behaviour_time',
                   'post_behaviour_time',
                   'bin_size', 'start_time', 'end_time']
        for column in columns:
            self.table_treeview.heading(column, text=column.capitalize().replace('_', ' '),
                                        command=lambda _col=column: self.treeview_sort_column(self.table_treeview, _col, False))
        for column in columns:
            # Set the width based on the column title length
            if tkf.Font().measure(column) > 110:
                self.table_treeview.column(
                    column, width=tkf.Font().measure(column) - 20)
            else:
                self.table_treeview.column(
                    column, width=(tkf.Font().measure(column)))

        table_hscrollbar.configure(command=self.table_canvas.xview)
        self.table_vscrollbar.configure(command=self.table_treeview.yview)

        # Configure row and column weights to make the frame expand
        table_container_frame.grid_rowconfigure(0, weight=1)
        table_container_frame.grid_columnconfigure(0, weight=1)

        # Bind the column click event to the callback function
        self.table_treeview.bind("<<TreeviewSelect>>",
                                 lambda event: self.on_row_click(event))

    def pick_sem_color(self):
        """Let the user choose a color for the SEM."""
        color_code = colorchooser.askcolor(title="Choose SEM colour")
        self.settings_manager.selected_sem_color = color_code[1]

        # Call handle_figure_display_selection function
        self.handle_figure_display_selection(event=None)

    def pick_trace_color(self):
        """Let the user choose a color for the trace."""
        # Open the color chooser dialog with the default color
        color = colorchooser.askcolor(
            color=self.settings_manager.selected_trace_color)

        if color[1] is not None:  # Check if a color was selected
            self.settings_manager.selected_trace_color = color[1]
        else:
            pass

        # Call handle_figure_display_selection function
        self.handle_figure_display_selection(event=None)

    def assign_behaviour_colors(self, behaviours):
        """
        Assign colors to the behaviours.

        Parameters:
        - behaviours: List of behaviours to assign colors to.
        """
        self.unique_behaviours = list(set(behaviours))

        for behaviour in self.unique_behaviours:
            if behaviour not in self.settings_manager.behaviour_colors:
                # New color generation
                self.behaviour_colors[behaviour] = self.graph_settings_container_instance.string_to_color(
                    behaviour)
            else:
                # Color from settings
                self.behaviour_colors[behaviour] = self.settings_manager.behaviour_colors[behaviour]

        # Create or update the BooleanVars for behaviour display status
        for behaviour in self.unique_behaviours:
            if behaviour not in self.behaviour_display_status:
                self.behaviour_display_status[behaviour] = tk.BooleanVar(
                    value=True)
            else:
                self.behaviour_display_status[behaviour].set(True)

        # Save the new color selection
        self.settings_manager.save_variables()

    # class HandleFiles:
    def save_static_inputs(self, df):
        """
        Save the static inputs to a JSON file.

        Parameters:
        - df: DataFrame containing behaviour settings to be saved.
        """

        behaviour_data = {}
        for _, row in df.iterrows():
            behaviour_name = row['Behaviour Name']
            pre_behaviour_time = row['Pre Behaviour Time']
            post_behaviour_time = row['Post Behaviour Time']
            bin_size = row['Bin Size']
            behaviour_data[behaviour_name] = (
                pre_behaviour_time, post_behaviour_time, bin_size)

        # Save the settings to a new JSON file
        with open('behaviour_settings.json', 'w') as f:
            json.dump(behaviour_data, f)

    def load_static_inputs(self):
        """Load the static inputs from a JSON file."""
        behaviour_settings = {}

        # Check if the JSON file exists
        if os.path.exists('behaviour_settings.json'):
            # If it does, open it and load the settings
            with open('behaviour_settings.json', 'r') as f:
                behaviour_data = json.load(f)

            # Convert the behaviour_data to the desired format
            for behaviour_name, settings in behaviour_data.items():
                pre_behaviour_time = settings[0]
                post_behaviour_time = settings[1]
                bin_size = settings[2]
                behaviour_settings[behaviour_name] = {'pre_behaviour_time': pre_behaviour_time,
                                                      'post_behaviour_time': post_behaviour_time,
                                                      'bin_size': bin_size}

        return behaviour_settings

    def select_event_file(self):
        """Select a behaviour file and parse it."""
        if self.checkbox_state and not self.data_selection_frame.baseline_button_pressed:
            messagebox.showinfo(
                "Error", "Please remember to save the baseline values.")
            return

        # Only clear previous selections if a file has been parsed before
        if self.is_file_parsed:
            self.clear_previous_behaviour_selections()

        manual_file_path = filedialog.askopenfilename(
            filetypes=[("CSV Files", "*.csv")])
        if manual_file_path:
            self.parse_manual_data(manual_file_path)
            self.is_file_parsed = True  # Set the flag to indicate that a file has been parsed

    def clear_previous_behaviour_selections(self):
        """Clear previous selections before parsing new event time file."""
        self.adjusted_behaviour_dataframes = {}
        self.original_start_times_min = None
        self.original_end_times_min = None
        self.unique_behaviours = []
        self.bar_items = []
        self.mean_duration = None
        self.sem_duration = None
        self.mean_sem_df = None
        self.current_table_key = None
        self.clear_table()
        self.table_treeview.delete(*self.table_treeview.get_children())

        # Clear the dropdown menu by removing all menu items
        menu = self.graph_settings_container_instance.behaviour_to_zero_dropdown["menu"]
        menu.delete(0, "end")

        # Disable the dropdown
        self.graph_settings_container_instance.behaviour_to_zero_dropdown['state'] = 'disabled'

        # Update to full trace display by default
        self.figure_display_dropdown.set("Full Trace Display")

        # Reset the selected value for behaviour_choice_graph to an empty string
        self.selected_behaviour.set("")

        # Disable the behaviour_choice_graph Combobox
        self.behaviour_choice_graph.configure(state=tk.DISABLED)

        # Clear behavior data and related visual elements
        self.create_behaviour_options(
            destroy_frame=True)  # Clear behavior options

        # Redraw the plot or refresh the app display
        self.handle_figure_display_selection(None)

    def create_behaviour_options(self, destroy_frame=True):
        """
        Create the behaviour options frame.

        Parameters:
        - destroy_frame: Boolean flag to indicate whether to destroy the existing frame.
        """
        if destroy_frame:
            self.destroy_existing_frame()

        self.behaviour_checkboxes = {}
        sorted_behaviours = sorted(self.unique_behaviours)
        self.settings_manager.update_unique_behaviours(self.unique_behaviours)

        self.graph_settings_container_instance.setup_canvas()
        self.create_control_buttons()
        self.create_behaviour_labels_and_controls(sorted_behaviours)

    def destroy_existing_frame(self):
        """Destroy the existing behaviour frame if it already exists."""
        if hasattr(self, "behaviour_frame"):
            self.behaviour_frame.destroy()

    def create_control_buttons(self):
        """Create control buttons for 'Select All' and 'Deselect All'."""
        select_all_button = tk.Button(
            self.graph_settings_container_instance.behaviour_frame,
            text='Select All',
            command=self.select_all
        )
        select_all_button.grid(row=0, column=1, padx=10,
                               pady=(5, 2), sticky=tk.W)

        deselect_all_button = tk.Button(
            self.graph_settings_container_instance.behaviour_frame,
            text='Deselect All',
            command=self.deselect_all
        )
        deselect_all_button.grid(
            row=0, column=2, padx=10, pady=(5, 2), sticky=tk.W)

    def create_behaviour_labels_and_controls(self, sorted_behaviours):
        """
        Create labels and control elements for behaviours.

        Parameters:
        - sorted_behaviours: List of behaviours sorted in alphabetical order.
        """
        behaviour_list_label = tk.Label(
            self.graph_settings_container_instance.behaviour_frame,
            text='Behaviours',
            bg='snow',
            font=('Helvetica', 12, 'bold')
        )
        behaviour_list_label.grid(
            row=0, column=0, padx=10, pady=(5, 2), sticky=tk.W)

        self.color_buttons = {}
        for i, behaviour in enumerate(sorted_behaviours, start=1):
            self.create_behaviour_control(behaviour, i)

    def create_behaviour_control(self, behaviour, row_index):
        """
        Create control elements for a single behaviour.

        Parameters:
        - behaviour: The behaviour to create controls for.
        - row_index: The row index to place the controls in.

        Returns:
        - None: If the behaviour is not found in the behaviour colors dictionary.
        """
        behaviour_color = self.behaviour_colors.get(behaviour)
        if behaviour_color is None:
            return

        color_code, text_color = self.get_color_code_and_text_color(
            behaviour_color)
        self.create_color_button(behaviour, color_code, text_color, row_index)
        self.create_checkbox(behaviour, row_index)

    def get_color_code_and_text_color(self, behaviour_color):
        """
        Get color code and appropriate text color based on brightness.

        Parameters:
        - behaviour_color: The color of the behaviour.

        Returns:
        - color_code: The hexadecimal color code.
        - text_color: The color of the text.
        """
        r, g, b = [int(c * 255) for c in behaviour_color[:3]]
        color_code = '#%02x%02x%02x' % (r, g, b)
        color_brightness = self.graph_settings_container_instance.brightness(
            behaviour_color)
        text_color = "black" if color_brightness > 0.5 else "white"
        return color_code, text_color

    def create_color_button(self, behaviour, color_code, text_color, row_index):
        """
        Create a button for selecting behavior color.

        Parameters:
        - behaviour: The behaviour to create a color button for.
        - color_code: The hexadecimal color code.
        - text_color: The color of the text.
        - row_index: The row index to place the button in.
        """
        color_button = tk.Button(
            self.graph_settings_container_instance.behaviour_frame,
            text=behaviour,
            command=lambda b=behaviour: self.graph_settings_container_instance.choose_color(
                b, color_button)
        )
        color_button.grid(row=row_index, column=0, padx=10,
                          pady=(5, 2), sticky=tk.W)
        color_button.config(fg=text_color, bg=color_code)
        self.color_buttons[behaviour] = color_button

    def create_checkbox(self, behaviour, row_index):
        """
        Create a checkbox for a single behaviour.

        Parameters:
        - behaviour: The behaviour to create a checkbox for.
        - row_index: The row index to place the checkbox in.
        """
        behaviour_option_checkbox = tk.Checkbutton(
            self.graph_settings_container_instance.behaviour_frame,
            variable=self.behaviour_display_status[behaviour],
            command=self.refresh_behaviour_options,
            bg='snow'
        )
        behaviour_option_checkbox.grid(
            row=row_index, column=1, padx=10, pady=(5, 2), sticky=tk.W)
        behaviour_option_checkbox.config(font=("Helvetica", 8))
        self.behaviour_checkboxes[behaviour] = behaviour_option_checkbox

    def update_box_colors_and_behaviour_options(self, behaviour, color_rgb):
        """
        Update the colors of the boxes with the specified behaviour on the graph and update the behaviour options.

        Parameters:
        - behaviour: The behaviour to update.
        - color_rgb: The new color for the behaviour, as an RGB tuple.
        """
        if behaviour in self.behaviour_boxes:
            for box in self.behaviour_boxes[behaviour]:
                box.set_facecolor(color_rgb)

            # Redraw only the color changes without affecting the rest of the graph display
            self.figure_canvas.draw_idle()

            # Update button color
            color_hex = '#%02x%02x%02x' % (
                int(color_rgb[0] * 255), int(color_rgb[1] * 255), int(color_rgb[2] * 255))
            # Retrieve the button reference from the dictionary
            button_to_update = self.color_buttons.get(behaviour)

            # Update the behaviour color in the dictionary
            self.behaviour_colors[behaviour] = color_rgb

            if button_to_update:
                button_to_update.config(bg=color_hex)

                behaviour_color = self.behaviour_colors[behaviour]
                # Calculate the brightness of the color
                color_brightness = self.graph_settings_container_instance.brightness(
                    behaviour_color)

                # Decide on text colour
                if color_brightness > 0.5:
                    text_color = "black"
                else:
                    text_color = "white"

                button_to_update.config(fg=text_color)

        # Update the behaviour colours for saving
        self.settings_manager.update_behaviour_colors(self.behaviour_colors)

    def parse_manual_data(self, file_path):
        """
        Parse a manually selected behaviour data file.

        Parameters:
        - file_path: Path to the CSV file to be parsed.

        Raises:
        - ValueError: If the required columns are not present in the CSV file.
        - ValueError: If the 'Behaviours/events' column is not present in the CSV file.
        """
        try:
            # Preliminary checks and settings
            if self.checkbox_state and not self.data_selection_frame.baseline_button_pressed:
                messagebox.showinfo(
                    "Error", "Please remember to save the baseline values.")
                return

            behaviour_settings = self.load_static_inputs()
            self.clear_table()

            # Read and process the file
            df, column_names = self.read_and_process_file(file_path)

            # Process each row of the DataFrame
            table_data, behaviour_names, behavior_durations = self.process_each_row(
                df, column_names, behaviour_settings, file_path)

            # Calculate the duration metrics for each behavior and store them in the cache
            self.calculate_and_store_behavior_metrics(behavior_durations)

            # Update the UI with the parsed data
            self.update_ui_with_manual_data(
                table_data, behaviour_names)

        except IOError:
            traceback.print_exc()
            messagebox.showerror("Error", "Failed to open the CSV file.")

        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to parse the CSV file: {e}")

    def read_and_process_file(self, file_path):
        """Read the CSV file and process its contents."""
        df = pd.read_csv(file_path)
        df.columns = map(str.lower, df.columns)
        column_names = self.prompt_column_names()
        column_names = {key: value.lower()
                        for key, value in column_names.items()}

        time_unit = self.time_input_unit_var.get().lower()
        if time_unit == "minutes":
            df[column_names['Start Time']] *= 60
            df[column_names['End Time']] *= 60

        required_columns = [column_names['Behaviours/events'],
                            column_names['Start Time'], column_names['End Time']]
        required_columns = [col.lower() for col in required_columns]

        behaviour_events_column = column_names['Behaviours/events'].lower()
        df = df.dropna(subset=[behaviour_events_column])

        for col in required_columns:
            if col not in df.columns:
                # Filter out columns containing 'Unnamed' and join the names
                available_columns = [c for c in df.columns if not c.lower().startswith('unnamed') and c not in required_columns]
                available_columns_str = ", ".join(available_columns)
                raise ValueError(
                    f'''Failed to parse the CSV file: Required column **'{col}'** is not present in the CSV file.\n\n'''
                    f'''Available columns (excluding already specified ones):\n  - {available_columns_str.replace(', ', '\n  - ')}'''
                )

        return df, column_names

    def process_each_row(self, df, column_names, behaviour_settings, file_path):
        """
        Process each row of the DataFrame and collect behavior data.

        Parameters:
        - df: DataFrame containing the behavior data.
        - column_names: Dictionary containing the column names.
        - behaviour_settings: Dictionary containing the behavior settings.
        - file_path: Path to the CSV file.

        Returns:
        - table_data: List of tuples containing the processed data.
        - behaviour_names: Set of unique behavior names.
        - behavior_durations: Dictionary containing behavior durations.
        """
        behaviour_names = set()
        table_data = []
        behavior_durations = {}
        default_pre_behaviour_time = "10"
        default_post_behaviour_time = "10"
        default_bin_size = "1"
        synchronize_start_time = float(
            self.behaviour_event_input_frame.synchronize_start_time_entry.get())

        for index, row in df.iterrows():
            behaviour_name = self.get_behaviour_name(
                row, column_names).capitalize()
            start_time, end_time = self.get_start_end_times(
                row, column_names, synchronize_start_time)
            behavior_durations = self.update_behavior_durations(
                behavior_durations, behaviour_name, start_time, end_time)
            pre_behaviour_time, post_behaviour_time, bin_size = self.get_behaviour_settings(
                behaviour_name, behaviour_settings, default_pre_behaviour_time, default_post_behaviour_time, default_bin_size)
            table_data.append((file_path, self.selected_column_var.get(), behaviour_name, self.get_behaviour_type(
                end_time), pre_behaviour_time, post_behaviour_time, bin_size, start_time, end_time))
            behaviour_names.add(behaviour_name)

        return table_data, behaviour_names, behavior_durations

    def get_behaviour_name(self, row, column_names):
        """
        Get and clean the behavior name from a row.

        Parameters:
        - row: The row to extract the behavior name from.
        - column_names: Dictionary containing the column names.

        Returns:
        - behaviour_name: The cleaned behavior name.
        """
        behaviour_name_raw = row[column_names['Behaviours/events']]
        if isinstance(behaviour_name_raw, float) and behaviour_name_raw.is_integer():
            behaviour_name = str(int(behaviour_name_raw))
        else:
            behaviour_name = str(behaviour_name_raw)
        return behaviour_name.strip()

    def get_start_end_times(self, row, column_names, synchronize_start_time):
        """
        Get the start and end times from a row.

        Parameters:
        - row: The row to extract the times from.
        - column_names: Dictionary containing the column names.
        - synchronize_start_time: The start time to synchronize with.

        Returns:
        - start_time: The start time.
        - end_time: The end time.
        """
        start_time = row[column_names['Start Time']] + synchronize_start_time
        end_time = row[column_names['End Time']] + synchronize_start_time
        return start_time, end_time

    def update_behavior_durations(self, behavior_durations, behaviour_name, start_time, end_time):
        """
        Update the behavior durations dictionary with new times.

        Parameters:
        - behavior_durations: Dictionary containing behavior durations.
        - behaviour_name: The name of the behavior.
        - start_time: The start time of the behavior.
        - end_time: The end time of the behavior.

        Returns:
        - behavior_durations: Updated dictionary containing behavior durations.
        """
        if behaviour_name not in behavior_durations:
            behavior_durations[behaviour_name] = {
                'start_times': [], 'end_times': []}
        behavior_durations[behaviour_name]['start_times'].append(start_time)
        behavior_durations[behaviour_name]['end_times'].append(end_time)
        return behavior_durations

    def get_behaviour_settings(self, behaviour_name, behaviour_settings, default_pre_behaviour_time, default_post_behaviour_time, default_bin_size):
        """
        Get the behavior settings, using stored settings if available.

        Parameters:
        - behaviour_name: The name of the behavior.
        - behaviour_settings: Dictionary containing the behavior settings.
        - default_pre_behaviour_time: The default pre-behavior time.
        - default_post_behaviour_time: The default post-behavior time.
        - default_bin_size: The default bin size.

        Returns:
        - pre_behaviour_time: The pre-behavior time.
        - post_behaviour_time: The post-behavior time.
        - bin_size: The bin size.
        """
        if behaviour_name in behaviour_settings:
            settings = behaviour_settings[behaviour_name]
            stored_pre_behaviour_time = settings.get('pre_behaviour_time', "")
            stored_post_behaviour_time = settings.get(
                'post_behaviour_time', "")
            stored_bin_size = settings.get('bin_size', "")
            if all(str(value) != "" for value in [stored_pre_behaviour_time, stored_post_behaviour_time, stored_bin_size]):
                return stored_pre_behaviour_time, stored_post_behaviour_time, stored_bin_size
        return default_pre_behaviour_time, default_post_behaviour_time, default_bin_size

    def get_behaviour_type(self, end_time):
        """
        Determine the behavior type based on the end time.

        Parameters:
        - end_time: The end time of the behavior.

        Returns:
        - The behavior type as either 'Point' or 'Continuous'.
        """
        if pd.isnull(end_time):
            return "Point"
        return "Continuous"

    def calculate_and_store_behavior_metrics(self, behavior_durations):
        """
        Calculate and store behavior duration metrics in the cache.

        Parameters:
        - behavior_durations: Dictionary containing behavior durations.
        """
        for behavior, times in behavior_durations.items():
            mean_duration, sem_duration = self.calculate_duration_metrics(
                times['start_times'], times['end_times'])
            cache_duration_key = behavior
            mean_duration_converted = self.convert_and_retrieve_time(
                mean_duration)
            sem_duration_converted = self.convert_and_retrieve_time(
                sem_duration)
            self.duration_data_cache[cache_duration_key] = {
                "mean_duration": mean_duration_converted,
                "sem_duration": sem_duration_converted,
                "mean_sem_df": None,  # Placeholder, update as needed
                "number_of_instances": len(times['start_times'])
            }

    def update_ui_with_manual_data(self, table_data, behaviour_names):
        """
        Update the UI with the processed manual data.

        Parameters:
        - table_data: List of tuples containing the processed data.
        - behaviour_names: Set of unique behavior names.
        """
        # DataFrame for display purposes only
        display_df = pd.DataFrame(table_data, columns=[
            "File Path", "Selected Column", "Behaviour Name", "Behaviour Type",
            "Pre Behaviour Time", "Post Behaviour Time", "Bin Size",
            "Start Time", "End Time"
        ])

        # Assign colours to behaviours
        self.assign_behaviour_colors(behaviour_names)

        # Update dropdowns and create behaviour options
        self.update_behaviour_dropdowns(behaviour_names, table_data)
        self.create_behaviour_options()

        # Generate a unique key for the table and store it
        self.current_table_key = str(uuid.uuid4())
        self.tables[self.current_table_key] = display_df

        # Update the table with new data
        self.update_table(display_df, new=True)

        # Save the static inputs
        self.save_static_inputs(display_df)

        # Keep a copy of the original table if it doesn't exist
        if not hasattr(self, 'original_table') or self.original_table is None:
            self.original_table = self.tables[self.current_table_key].copy()

        # Handle figure display selection
        self.handle_figure_display_selection(None)

    def update_behaviour_dropdowns(self, behaviour_names, table_data):
        """
        Update the dropdown menus with the behaviour names.

        Parameters:
        - behaviour_names: Set of unique behaviour names.
        - table_data: List of tuples containing the processed data.
        """
        sorted_behaviours = sorted(list(behaviour_names))

        # Update static_inputs_frame behaviour dropdown menu
        self.update_menu(self.static_inputs_frame.behaviour_dropdown, [
                         'All Behaviours'] + sorted_behaviours, self.static_inputs_frame.selected_behaviour)

        # Update graph settings container behaviour dropdown menu
        single_instance_behaviours = self.get_single_instance_behaviours(
            table_data)
        self.update_menu(self.graph_settings_container_instance.behaviour_to_zero_dropdown,
                         single_instance_behaviours, self.graph_settings_container_instance.selected_behaviour_to_zero)

        if single_instance_behaviours:
            first_behaviour = single_instance_behaviours[0]
            self.graph_settings_container_instance.selected_behaviour_to_zero.set(
                first_behaviour)
            self.graph_settings_container_instance.selected_behaviour_to_zero.trace_add(
                'write', self.handle_behaviour_change)

        # Enable dropdowns
        self.behaviour_choice_graph["values"] = sorted_behaviours
        self.behaviour_choice_graph['state'] = 'normal'
        self.graph_settings_container_instance.behaviour_to_zero_dropdown['state'] = 'normal'
        self.static_inputs_frame.behaviour_dropdown['state'] = 'normal'

    def update_menu(self, menu_widget, choices, variable):
        """Helper function to update a dropdown menu.

        Parameters:
        - menu_widget: The dropdown menu widget.
        - choices: List of choices to populate the dropdown menu.
        - variable: The variable associated with the dropdown menu.
        """
        menu = menu_widget["menu"]
        menu.delete(0, "end")
        for choice in choices:
            menu.add_command(label=choice, command=tk._setit(variable, choice))

    def get_single_instance_behaviours(self, table_data):
        """
        Get behaviours with a single instance from the table data.

        Parameters:
        - table_data: List of tuples containing the processed data.

        Returns:
        - List of behaviours with a single instance.
        """
        behaviour_counts = {}
        for data_tuple in table_data:
            behaviour_name = data_tuple[2]
            behaviour_counts[behaviour_name] = behaviour_counts.get(
                behaviour_name, 0) + 1

        return [behaviour for behaviour, count in behaviour_counts.items() if count == 1]

    # class UserInputHandling:

    def calculate_duration_metrics(self, start_times, end_times):
        """
        Calculate the mean and SEM of durations given start and end times.

        Parameters:
        - start_times (List[float]): List of start times.
        - end_times (List[float]): List of end times.

        Returns:
        - mean_duration (float): Mean duration of the events.
        - sem_duration (float): Standard error of the mean (SEM) of the durations.
        """

        durations = [(end - start) / 60 for start,
                     end in zip(start_times, end_times) if end is not None]

        if not durations:
            return np.nan, np.nan  # Return NaN if there are no valid durations

        mean_duration = np.nanmean(durations)
        sem_duration = np.nanstd(
            durations) / np.sqrt(len(durations) - np.isnan(durations).sum())
        return mean_duration, sem_duration

    def on_row_click(self, event):
        """
        Handle row selection in the table treeview.

        Parameters:
        - event: The event that triggered the row click.
        """
        selected_row = self.table_treeview.focus()
        if not selected_row:
            return
        values = self.table_treeview.item(selected_row)['values']

        behaviour_name = values[2]

        # Get the current dataframe associated with the current_table_key
        current_df = self.tables[self.current_table_key]

        # Filter the dataframe for the selected behaviour name
        selected_behaviour_df = current_df[current_df['Behaviour Name']
                                           == behaviour_name]

        if not selected_behaviour_df.empty:
            self.start_time = float(
                selected_behaviour_df.iloc[0]['Start Time']) / 60
            self.end_time = float(
                selected_behaviour_df.iloc[0]['End Time']) / 60
            self.pre_behaviour_time = float(
                selected_behaviour_df.iloc[0]['Pre Behaviour Time']) / 60
            self.post_behaviour_time = float(
                selected_behaviour_df.iloc[0]['Post Behaviour Time']) / 60
            self.column_used = selected_behaviour_df.iloc[0]['Selected Column']

        self.figure_display_dropdown.set("Single Row Display")
        self.handle_figure_display_selection(event=None)

    def handle_new_data_file(self, file_path_var, selected_column_var, column_dropdown, mouse_name, dataframe, *args):
        """
        Handle loading and initializing a new data file.

        Parameters:
        - file_path_var: Variable for the file path.
        - selected_column_var: Variable for the selected column.
        - column_dropdown: Dropdown menu for column selection.
        - mouse_name: Name of the mouse.
        - dataframe: DataFrame containing the data.
        - *args: Additional arguments.
        """
        self.file_path_var = file_path_var
        self.selected_column_var = selected_column_var
        self.column_dropdown = column_dropdown
        self.mouse_name = mouse_name
        self.dataframe = dataframe
        self.baseline_button_pressed = False
        self.z_score_computed = False
        self.baseline_data_mean = None
        self.baseline_data_std = None
        self.data_already_adjusted = False
        self.first_offset_time = None
        self.warning_shown = False

        self.graph_settings_container_instance.zero_x_axis_checkbox_var.set(
            0)  # Uncheck the checkbox

        if 'dFoF_465' in dataframe.columns:
            self.selected_column_var.set('dFoF_465')
        else:
            self.selected_column_var.set(dataframe.columns[1])

        self.selected_column_var.trace(
            'w', lambda *args: self.data_selection_frame.on_column_selection_changed())

        self.unique_behaviours = []
        self.clear_table()
        self.table_treeview.delete(*self.table_treeview.get_children())
        # Clear the dictionary of adjusted behaviour dataframes
        self.adjusted_behaviour_dataframes = {}

        # Update to full trace display by default
        self.figure_display_dropdown.set("Full Trace Display")

        # Reset the selected value for behaviour_choice_graph to an empty string
        self.selected_behaviour.set("")

        # Disable the behaviour_choice_graph Combobox
        self.behaviour_choice_graph.configure(state=tk.DISABLED)

        # Clear behavior data and related visual elements
        self.create_behaviour_options(
            destroy_frame=True)  # Clear behavior options

        # If 'ax' is defined, clear its content
        if hasattr(self, 'ax'):
            self.ax.clear()  # Clear the plot

        # Clear the dropdown menu by removing all menu items
        menu = self.graph_settings_container_instance.behaviour_to_zero_dropdown["menu"]
        menu.delete(0, "end")

        # Reset the selected value to an empty string or a default value
        self.graph_settings_container_instance.selected_behaviour_to_zero.set(
            "")  # or set it to a default value if needed

        # Disable the dropdown
        self.graph_settings_container_instance.behaviour_to_zero_dropdown['state'] = 'disabled'

        self.handle_figure_display_selection(None)

    def save_inputs(self):
        """
        Function to save the user inputs to a JSON file.

        Returns:
        - None: If the user inputs are not valid.
        """
        pre_behaviour_time = self.static_inputs_frame.pre_behaviour_time_entry.get()
        post_behaviour_time = self.static_inputs_frame.post_behaviour_time_entry.get()
        bin_size = self.static_inputs_frame.bin_size_entry.get()

        selected_behaviour = self.static_inputs_frame.selected_behaviour.get()

        # Get the current dataframe associated with the current_table_key
        current_df = self.tables.get(self.current_table_key)

        if current_df is None:
            print("No current dataframe found.")
            return

        # Update the values in the current dataframe
        if selected_behaviour == 'All Behaviours':
            self.update_behaviour_times(
                current_df, pre_behaviour_time, post_behaviour_time, bin_size)
        else:
            self.update_selected_behaviour_times(
                current_df, selected_behaviour, pre_behaviour_time, post_behaviour_time, bin_size)

        # Update the corresponding dataframe in self.tables
        self.tables[self.current_table_key] = current_df

        # Update the table display
        self.update_table(current_df, new=True)

        self.save_static_inputs(current_df)

    def update_behaviour_times(self, df, pre_time, post_time, bin_size):
        """
        Update the pre and post behaviour times and bin size for all behaviours.

        Parameters:
        - df: DataFrame containing behaviour settings.
        - pre_time: The pre-behaviour time.
        - post_time: The post-behaviour time.
        - bin_size: The bin size.
        """
        df['Pre Behaviour Time'] = pre_time
        df['Post Behaviour Time'] = post_time
        df['Bin Size'] = bin_size

    def update_selected_behaviour_times(self, df, behaviour, pre_time, post_time, bin_size):
        """
        Update the pre and post behaviour times and bin size for a specific behaviour.

        Parameters:
        - df: DataFrame containing behaviour settings.
        - behaviour: The behaviour to update.
        - pre_time: The pre-behaviour time.
        - post_time: The post-behaviour time.
        - bin_size: The bin size.
        """
        mask = df['Behaviour Name'] == behaviour
        df.loc[mask, 'Pre Behaviour Time'] = pre_time
        df.loc[mask, 'Post Behaviour Time'] = post_time
        df.loc[mask, 'Bin Size'] = bin_size

    def toggle_behaviour_start_time(self):
        """Function to toggle the behaviour start time entry based on the behaviour type."""
        if self.behaviour_event_input_frame.behaviour_type_var.get() == "Point":
            self.behaviour_event_input_frame.start_time_entry.config(
                state=tk.NORMAL)
            self.behaviour_event_input_frame.end_time_entry.config(
                state=tk.DISABLED)
            self.behaviour_event_input_frame.end_time_var.set(
                "")  # Reset the end time variable
        else:
            self.behaviour_event_input_frame.start_time_entry.config(
                state=tk.NORMAL)
            self.behaviour_event_input_frame.end_time_entry.config(
                state=tk.NORMAL)

    def toggle_behaviour_end_time(self):
        """Function to toggle the behaviour end time entry based on the behaviour type."""
        if self.behaviour_event_input_frame.behaviour_type_var.get() == "Continuous":
            self.behaviour_event_input_frame.start_time_entry.config(
                state=tk.NORMAL)
            self.behaviour_event_input_frame.end_time_entry.config(
                state=tk.NORMAL)
        else:
            self.behaviour_event_input_frame.start_time_entry.config(
                state=tk.NORMAL)
            self.behaviour_event_input_frame.end_time_entry.config(
                state=tk.DISABLED)

    def select_all(self):
        """Function to select all behaviours."""
        for behaviour, checkbox in self.behaviour_checkboxes.items():
            checkbox.select()
            self.behaviour_display_status[behaviour].set(1)
        self.refresh_behaviour_options()

    def deselect_all(self):
        """Function to deselect all behaviours."""
        for behaviour, checkbox in self.behaviour_checkboxes.items():
            checkbox.deselect()
            self.behaviour_display_status[behaviour].set(0)
        self.refresh_behaviour_options()

    def refresh_behaviour_options(self):
        """Refresh the behaviour options"""
        for behaviour in self.behaviour_boxes:
            if not self.behaviour_display_status[behaviour].get():
                # Hide all the boxes associated with the behaviour
                for box in self.behaviour_boxes[behaviour]:
                    box.set_visible(False)
            else:
                # Show all the boxes associated with the behaviour
                for box in self.behaviour_boxes[behaviour]:
                    box.set_visible(True)

        self.figure_canvas.draw_idle()

    def extract_button_click_handler(self):
        """Handle the extract button click."""
        params = self.check_and_prepare_parameters(self.current_table_key)

        if params is None:
            return

        behaviours_to_export = {
            behaviour for behaviour in params["behaviours_to_export"] if self.behaviour_display_status[behaviour].get() == 1}

        if behaviours_to_export:
            behaviours_str = ", ".join(behaviours_to_export)
            proceed_message = f"The following behaviours will be exported: {
                behaviours_str}\nDo you want to proceed?"
            proceed = messagebox.askokcancel(
                "Behaviours to Export", proceed_message)
        else:
            proceed_message = "No behaviours have been selected for export.\nDo you want to proceed?"
            proceed = messagebox.askokcancel(
                "No Behaviours Selected", proceed_message)

        if proceed:
            file_path = self.data_selection_frame.file_path_var.get()
            self.extract_data_from_photometry(file_path, params)
            print("Data extraction complete.")
        else:
            print("Data extraction cancelled.")

        # Save the new color selection regardless of user decision
        self.settings_manager.save_variables()

    def select_column_names(self):
        """Function to select the column names in the CSV file."""
        # Load column names from JSON file
        current_column_names = self.load_column_names()

        # Create the settings window
        settings_window = tk.Toplevel(self.winfo_toplevel())
        settings_window.title("Column Names in CSV")

        # Create labels and entry boxes for each column name
        column_labels = ["Behaviours/events", "Start Time", "End Time"]
        entry_boxes = self.create_column_entries(
            settings_window, column_labels, current_column_names)

        # Dropdown menu for time unit selection
        self.create_time_unit_dropdown(
            settings_window, len(column_labels), current_column_names)

        # Create the save button
        save_button = tk.Button(settings_window, text="Save", command=lambda: self.save_column_names(
            settings_window, entry_boxes))
        save_button.grid(row=len(column_labels) + 1, column=0,
                         columnspan=2, padx=10, pady=5, sticky=tk.NSEW)

        # Update idle tasks to get updated dimensions and then center the window
        settings_window.update_idletasks()
        center_window_on_screen(settings_window)

    def load_column_names(self):
        """
        Load column names from JSON file.

        Returns:
        - column_names: Dictionary containing the column names.

        Exceptions:
        - IOError: If the file does not exist or cannot be read.
        - json.JSONDecodeError: If the file is not a valid JSON file.
        """
        try:
            with open("column_names.json", "r") as file:
                return json.load(file)
        except (IOError, json.JSONDecodeError):
            return self.default_column_names

    def create_column_entries(self, window, labels, current_names):
        """
        Create entry boxes for column names.

        Parameters:
        - window: The window to create the entry boxes in.
        - labels: List of column labels.
        - current_names: Dictionary containing the current column names.

        Returns:
        - entry_boxes: Dictionary of entry boxes containing column names.
        """
        entry_boxes = {}
        for i, label in enumerate(labels):
            tk.Label(window, text=label, font=('Helvetica', 10)).grid(
                row=i, column=0, padx=10, pady=5)
            entry_box = tk.Entry(window)
            entry_box.insert(0, current_names[label])
            entry_box.grid(row=i, column=1, padx=10, pady=5)
            entry_boxes[label] = entry_box
        return entry_boxes

    def create_time_unit_dropdown(self, window, row, current_names):
        """
        Create a dropdown menu for time unit selection.

        Parameters:
        - window: The window to create the dropdown menu in.
        - row: The row to place the dropdown menu in.
        - current_names: Dictionary containing the current column names.
        """
        tk.Label(window, text="Time Unit", font=('Helvetica', 10)).grid(
            row=row, column=0, padx=10, pady=5)
        time_unit_options = ["seconds", "minutes"]
        self.time_input_unit_var = tk.StringVar(window)
        self.time_input_unit_var.set(current_names.get("Time Unit", "seconds"))
        time_unit_dropdown = ttk.Combobox(
            window, state="readonly", values=time_unit_options, width=15, textvariable=self.time_input_unit_var)
        time_unit_dropdown.grid(row=row, column=1, padx=10, pady=5)

    def save_column_names(self, settings_window, entry_boxes):
        """
        Save the column names in a JSON file.

        Parameters:
        - settings_window: The settings window to be closed after saving.
        - entry_boxes: Dictionary of entry boxes containing column names.
        """
        column_names = {}

        # Get the values from the entry boxes
        for label, entry_box in entry_boxes.items():
            # Get the text from the box and strip leading/trailing spaces
            text = entry_box.get().strip()

            # Store the text in the column names dictionary
            column_names[label] = text

        # Save the selected time unit
        column_names["Time Unit"] = self.time_input_unit_var.get()

        # Assign to self.column_names
        self.column_names = column_names

        # Save the column names in the configuration file or wherever you prefer
        with open("column_names.json", "w") as file:
            json.dump(column_names, file)

        # Close the settings window
        settings_window.destroy()

    def prompt_column_names(self):
        """
        Prompt the user to enter the column names.

        Returns:
        - column_names: Dictionary containing the column names.

        Exceptions:
        - IOError: If the file does not exist or cannot be read.
        - json.JSONDecodeError: If the file is not a valid JSON file.
        """
        # Load column names from JSON file or return default values if file doesn't exist or is invalid
        try:
            with open("column_names.json", "r") as file:
                column_names = json.load(file)
        except (IOError, json.JSONDecodeError):
            column_names = self.default_column_names

            # Convert the column names dictionary to lowercase
        column_names = {key: value.lower()
                        for key, value in column_names.items()}

        # Set the time input unit variable based on the loaded value (default to 'seconds' if not present)
        self.time_input_unit_var.set(
            column_names.get("Time Unit", "seconds").lower())

        return column_names

    def truncate_sheet_title(self, title, max_length=31):
        """
        Truncate the sheet title to a maximum length.

        Parameters:
        - title: The original sheet title.
        - max_length: The maximum allowed length for the sheet title (default is 31 characters).

        Returns:
        - truncated_title: The truncated sheet title.
        """
        # Split the title into words
        words = title.split()

        # Truncate the longest word if necessary
        if len(title) > max_length:
            longest_word = max(words, key=len)
            if len(longest_word) > 3:
                truncated_word = longest_word[:3] + "..."
                words[words.index(longest_word)] = truncated_word

        # Reconstruct the title
        truncated_title = " ".join(words)
        truncated_title = re.sub(r'\s+', ' ', truncated_title.strip())

        # Ensure the title does not exceed max_length
        if len(truncated_title) > max_length:
            truncated_title = truncated_title[:max_length].strip()

        return truncated_title

    # class DataManipulation:
    def treeview_sort_column(self, tv, col, reverse):
        """
        Sort the treeview by the specified column.

        Parameters:
        - tv: The treeview widget.
        - col: The column to sort by.
        - reverse: Boolean indicating whether to sort in reverse order.
        """
        values = [(tv.set(k, col), k) for k in tv.get_children('')]

        def convert(val):
            """
            Convert a string value to a float, if possible.

            Parameters:
            - val: The value to convert.

            Returns:
            - The converted float value, or the original value if conversion is not possible.
            """
            if val == '':
                return None
            try:
                return float(val)
            except ValueError:
                return val

        values = [(convert(val), k) for val, k in values]

        if reverse:
            values.sort(key=lambda t: float('-inf')
                        if t[0] is None else t[0], reverse=True)
        else:
            values.sort(key=lambda t: float('inf') if t[0] is None else t[0])

        for index, (val, k) in enumerate(values):
            tv.move(k, '', index)

        tv.heading(col, command=lambda _col=col: self.treeview_sort_column(
            tv, _col, not reverse))

    def add_transparent_boxes(self, ax, data, behaviours, start_times_min, end_times_min, start_point=None, end_point=None):
        """
        Add transparent boxes to the plot.

        Parameters:
        - ax: The axis on which to add the boxes.
        - data: The data for plotting.
        - behaviours: List of behaviours to add boxes for.
        - start_times_min: List of start times for each behaviour.
        - end_times_min: List of end times for each behaviour.
        - start_point: Optional starting point for the x-axis range.
        - end_point: Optional ending point for the x-axis range.
        """
        time_unit = self.graph_settings_container_instance.time_unit_menu.get()
        time_factor = self.get_time_scale(time_unit)

        # Create a copy of the time data
        if self.checkbox_state:
            time_data = self.dataframe.get(
                'z_scored_time', self.calculate_z_score()[1]).copy()
        else:
            time_data = self.dataframe.iloc[:, 0].copy()

        time_data *= time_factor
        if start_point is not None:
            start_point *= time_factor
        if end_point is not None:
            end_point *= time_factor

        if start_point is not None and end_point is not None:
            filtered_data = data[(time_data >= start_point)
                                 & (time_data <= end_point)]
            min_y_value, max_y_value = filtered_data.min(), filtered_data.max()
            ax.set_xlim(start_point, end_point)
        else:
            min_y_value, max_y_value = data.min(), data.max()
            time_data_clean = time_data.dropna().reset_index(drop=True)
            if not time_data_clean.empty:
                ax.set_xlim(time_data_clean.iloc[0], time_data_clean.iloc[-1])
            else:
                print("Warning: time_data_clean is empty. Cannot set xlim.")

        for behaviour, start_time_min, end_time_min in zip(behaviours, start_times_min, end_times_min):
            if behaviour not in self.behaviour_display_status or not self.behaviour_display_status[behaviour].get() or not behaviour:
                continue

            min_box_width = 0.01  # Minimum box width in minutes
            x_position = start_time_min if start_time_min is not None else start_point
            end_x_position = end_time_min if end_time_min is not None else end_point
            box_width = max(
                end_x_position - x_position if end_x_position is not None else min_box_width, min_box_width)
            box_height = float(self.graph_settings_container_instance.box_height_entry.get(
            )) * ((max_y_value - min_y_value) * 1.05)
            box_y_position = min_y_value - abs(0.05 * min_y_value)
            alpha_value = min(
                max(float(self.graph_settings_container_instance.alpha_entry.get()), 0), 1)

            box = Rectangle((x_position * time_factor, box_y_position), box_width * time_factor,
                            box_height, facecolor=self.behaviour_colors[behaviour], alpha=alpha_value)

            if behaviour in self.behaviour_boxes:
                self.behaviour_boxes[behaviour].append(box)
            else:
                self.behaviour_boxes[behaviour] = [box]
            ax.add_patch(box)

    def clear_table(self):
        """Function to clear and cache the table."""
        self.original_table = None
        self.current_table_key = None
        self.tables.clear()
        self.duration_data_cache = {}

    def adjust_start_end_times(self, dataframe):
        """
        Adjust the start and end times based on checkbox state.

        Parameters:
        - dataframe: The DataFrame containing start and end times.

        Returns:
        - dataframe: The adjusted DataFrame.
        """
        baseline_offset = float(
            self.data_selection_frame.baseline_start_entry.get())
        dataframe['Start Time'] -= baseline_offset
        dataframe['End Time'] = pd.to_numeric(
            dataframe['End Time'], errors='coerce')
        dataframe['End Time'] = dataframe['End Time'].apply(
            lambda x: x - baseline_offset if pd.notnull(x) else x)

        return dataframe

    def update_table_from_frame(self):
        """Function to update the table from the DataFrame."""
        if self.current_table_key is not None and self.current_table_key in self.tables:
            current_df = self.tables[self.current_table_key].copy()
            self.update_table(current_df)
            self.handle_figure_display_selection(None)
            # Clear the dictionary of adjusted behaviour dataframes
            self.adjusted_behaviour_dataframes = {}

    def update_table(self, dataframe, new=False):
        """
        Update the table with the data from the DataFrame.

        Parameters:
        - dataframe: The DataFrame containing the data to display in the table.
        - new: Boolean flag indicating whether the table is being updated with new data.
        """
        if self.checkbox_state and new is not True:
            dataframe = self.adjust_start_end_times(dataframe)
            self.tables[self.current_table_key] = dataframe

        # Check for negative (Start Time - Pre Behaviour Time) and warn the user
        if not self.warning_shown:
            negative_behaviours = []
            for _, row in dataframe.iterrows():
                try:
                    start_time = float(row['Start Time'])
                    pre_behaviour_time = float(row['Pre Behaviour Time'])

                    if (start_time - pre_behaviour_time) < 0:
                        negative_behaviours.append(row['Behaviour Name'])
                except ValueError:
                    # Handle cases where conversion to float fails
                    continue

            if negative_behaviours:
                negative_behaviours_str = ", ".join(negative_behaviours)
                messagebox.showwarning(
                    "Negative Time Warning",
                    f"The following behaviours have a start time that, when adjusted by the pre-behaviour time, becomes negative: {
                        negative_behaviours_str}"
                )
                # Set the flag to indicate that the warning has been shown
                self.warning_shown = True

        # Filter out rows where Start Time is negative
        try:
            dataframe['Start Time'] = dataframe['Start Time'].astype(float)
        except ValueError:
            # Handle cases where conversion to float fails, you may want to log this or handle it differently
            pass

        dataframe = dataframe[dataframe['Start Time'] >= 0]

        self.table_treeview.delete(*self.table_treeview.get_children())
        self.populate_table(dataframe)
        self.adjust_column_widths()
        self.update_table_scrollbar()

    def populate_table(self, dataframe):
        """Add rows from the DataFrame to the table treeview."""
        for index, row in dataframe.iterrows():
            file_name = os.path.basename(row['File Path'])
            values = [file_name, row['Selected Column'], row['Behaviour Name'], row['Behaviour Type'],
                      row['Pre Behaviour Time'], row['Post Behaviour Time'], row['Bin Size'],
                      row['Start Time'], row['End Time']]
            self.table_treeview.insert('', 'end', values=values)

    def adjust_column_widths(self):
        """Adjust column widths based on content."""
        columns = ['file_name', 'column_title', 'behaviour_name', 'behaviour_type', 'pre_behaviour_time', 'post_behaviour_time',
                   'bin_size', 'start_time', 'end_time']
        for column in columns:
            self.table_treeview.column(
                column, width=tkf.Font().measure(column))

    def update_table_scrollbar(self):
        """Update the scrollbar settings for the table."""
        self.table_treeview.configure(yscrollcommand=self.table_vscrollbar.set)
        self.table_treeview.update_idletasks()

    def convert_time(self, unit, time_in_seconds):
        """
        Convert time from seconds to the specified unit.

        Parameters:
        - unit: The unit to convert to ('hours', 'minutes', or 'seconds').
        - time_in_seconds: The time value in seconds.

        Returns:
        - The time converted to the specified unit.
        """
        if unit == "hours":
            return time_in_seconds / 3600
        if unit == "minutes":
            return time_in_seconds / 60
        return time_in_seconds  # assumes time is already in seconds

    def calculate_mean_sem(self, dataframe, ignore_col):
        """
        Calculate the mean and SEM for each row in the DataFrame.

        Parameters:
        - dataframe: The DataFrame containing the data.
        - ignore_col: The column to ignore in the calculations.

        Returns:
        - calculated_mean_sem_df: A DataFrame with 'Time', 'Mean', and 'SEM' columns.
        """
        # Drop the ignore_col from the DataFrame
        data_without_ignore_col = dataframe.drop(ignore_col, axis=1)

        # If there's only one instance of the behavior
        if len(data_without_ignore_col.columns) == 1:
            single_value = data_without_ignore_col.iloc[:, 0]
            mean_sem_df = pd.DataFrame({
                'Time': dataframe['Time'],
                'Mean': single_value,  # Use the single behavior's value for Mean
                'SEM': [0] * len(single_value)  # Fill SEM with zeros
            })
            # Set calculated_mean_sem_df to mean_sem_df
            calculated_mean_sem_df = mean_sem_df

        else:
            # Calculate the mean for each row
            means = data_without_ignore_col.mean(axis=1)
            # Calculate the SEM for each row
            sems = data_without_ignore_col.sem(axis=1)
            # Create a DataFrame with the means and SEMs
            calculated_mean_sem_df = pd.DataFrame(
                {'Time': dataframe['Time'], 'Mean': means, 'SEM': sems})

        return calculated_mean_sem_df

    def generate_mean_sem_df(self, behaviour_data_by_instance, time_points, start_time_adjusted, end_time_adjusted):
        """
        Generate a DataFrame with the mean and SEM for each time point.

        Parameters:
        - behaviour_data_by_instance: Dictionary containing behaviour data by instance.
        - time_points: List of time points.
        - start_time_adjusted: Adjusted start time.
        - end_time_adjusted: Adjusted end time.

        Returns:
        - mean_sem_df: A DataFrame with 'Time', 'Mean', and 'SEM' columns.
        """
        # Create a DataFrame using the behaviour_data_by_instance dictionary
        behaviour_data_frame = pd.DataFrame(behaviour_data_by_instance)

        # Insert the time_points as a new column at the beginning of the DataFrame
        behaviour_data_frame.insert(0, 'Time', np.linspace(
            start_time_adjusted, end_time_adjusted, len(time_points)))

        # Calculate mean and SEM
        mean_sem_df = self.calculate_mean_sem(behaviour_data_frame, 'Time')

        return mean_sem_df

    def calculate_z_score(self):
        """
        TODO TEST THIS FUNCTION
        Calculate the z-score for the selected column.

        Returns:
        - z_scored_data: The z-scored data for the selected column.
        - z_scored_time: The corresponding z-scored time data.
        """
        selected_column = self.selected_column_var.get()
        if self.figure_display_dropdown.get() != "Z-scored data":
            self.figure_display_dropdown.set("Z-scored data")

        current_baseline_start = self._get_baseline_start() / 60
        current_baseline_end = self._get_baseline_end() / 60

        if self._is_z_score_computed(current_baseline_start, current_baseline_end):
            return self.dataframe['baselined_z_score'], self.dataframe['z_scored_time']

        baseline_start_time_min = current_baseline_start
        baseline_end_time_min = current_baseline_end

        baseline_data, raw_data = self._get_baseline_and_raw_data(
            baseline_start_time_min, baseline_end_time_min, selected_column)

        self.baseline_data_mean = np.mean(baseline_data)
        self.baseline_data_std = np.std(baseline_data)

        z_scored_data, z_scored_time = self._compute_z_scores(
            raw_data, baseline_start_time_min)

        min_length = min(len(z_scored_data), len(z_scored_time))
        z_scored_data = z_scored_data[:min_length]
        z_scored_time = z_scored_time[:min_length]

        self._update_dataframe(z_scored_data, z_scored_time)

        self._update_flags(current_baseline_start, current_baseline_end)

        return z_scored_data, z_scored_time

    def _get_baseline_start(self):
        return float(self.data_selection_frame.baseline_start_entry.get())

    def _get_baseline_end(self):
        return float(self.data_selection_frame.baseline_end_entry.get())

    def _is_z_score_computed(self, current_baseline_start, current_baseline_end):
        return (self.z_score_computed and
                self.previous_baseline_start == current_baseline_start and
                self.previous_baseline_end == current_baseline_end)

    def _get_baseline_and_raw_data(self, baseline_start_time_min, baseline_end_time_min, selected_column):
        baseline_data = self.dataframe.loc[
            (self.dataframe.iloc[:, 0] >= baseline_start_time_min) &
            (self.dataframe.iloc[:, 0] < baseline_end_time_min), selected_column].reset_index(drop=True)
        raw_data = self.dataframe.loc[
            (self.dataframe.iloc[:, 0] >= baseline_start_time_min), selected_column].reset_index(drop=True)
        return baseline_data, raw_data

    def _compute_z_scores(self, raw_data, baseline_start_time_min):
        z_scored_data = (raw_data - self.baseline_data_mean) / \
            self.baseline_data_std
        z_scored_time = self.dataframe.loc[
            self.dataframe.iloc[:, 0] >= baseline_start_time_min].iloc[:, 0].reset_index(drop=True)
        z_scored_time -= baseline_start_time_min
        return z_scored_data, z_scored_time

    def _update_dataframe(self, z_scored_data, z_scored_time):
        self.dataframe['baselined_z_score'] = z_scored_data
        self.dataframe['z_scored_time'] = z_scored_time

    def _update_flags(self, current_baseline_start, current_baseline_end):
        self.z_score_computed = True
        self.previous_baseline_start = current_baseline_start
        self.previous_baseline_end = current_baseline_end

    def calculate_auc(self, data, dx=0.1):
        """
        Calculate Area Under Curve (AUC) using Simpson's rule.

        Parameters:
        - data: List or array of data points.
        - dx: The spacing between data points (time step).

        Returns:
        - float: The computed area under the curve.
        """
        return simpson(data, dx=dx)

    def calculate_max_amp(self, data):
        """
        Calculate Maximum Amplitude.

        Parameters:
        - data: List or array of data points.

        Returns:
        - float: The maximum amplitude, or NaN if data is empty.
        """
        if len(data) == 0:
            return np.nan
        return np.max(data)

    def calculate_mean_dff(self, data):
        """
        Calculate Mean of Delta F / F (Mean dF/F).

        Parameters:
        - data: List or array of data points.

        Returns:
        - float: The mean dF/F.
        """
        return np.mean(data)

    def extract_duration_data(self):
        """
        Extract duration data from the cache and create a DataFrame.

        Returns:
        - df_duration: A DataFrame containing the behavior, mean duration, SEM, and number of instances.
        """
        duration_data = []

        for key, value in self.duration_data_cache.items():
            behavior = key  # Assuming the key is the behavior name
            mean_duration = value["mean_duration"] * 60
            sem_duration = value["sem_duration"] * 60
            number_of_instances = value["number_of_instances"]
            duration_data.append(
                [behavior, mean_duration, sem_duration, number_of_instances])

        behavior_duration_stats_df = pd.DataFrame(duration_data, columns=[
            'Behavior', 'Mean Duration (s)', 'SEM (s)', 'Number of Instances'])
        behavior_duration_stats_df = behavior_duration_stats_df.sort_values(
            by='Behavior')  # Sort by 'Behavior' column
        return behavior_duration_stats_df

    def create_extraction_folder(self, file_path):
        """
        Create a folder to store the extracted data.

        Parameters:
        - file_path: Path to the photometry data file.

        Returns:
        - folder_path: Path to the created folder.
        """
        folder_path = os.path.join(os.path.dirname(file_path), os.path.splitext(
            os.path.basename(file_path))[0] + "_NeuroBehaviorSync")
        os.makedirs(folder_path, exist_ok=True)
        return folder_path

    def handle_zscore_checkbox_export(self, behavior_duration_stats_df):
        """ 
        Handle the z-score checkbox state for data export.

        Parameters:
        - df_duration: DataFrame containing duration data.

        Returns:
        - z_scored_data: The z-scored data.
        - z_scored_time: The z-scored time data.
        - df_duration: Updated DataFrame with mean and STD data.
        """
        # Retrieve z-scored data and its associated time directly from the dataframe
        z_scored_data = self.dataframe['baselined_z_score']
        z_scored_time = self.dataframe['z_scored_time']

        # Append mean z score data to df_duration skipping a column with titles mean and sem
        behavior_duration_stats_df[''] = [
            np.nan] * (len(behavior_duration_stats_df))
        behavior_duration_stats_df['Mean df/f for baseline'] = [self.baseline_data_mean] + \
            [np.nan] * (len(behavior_duration_stats_df) - 1)
        behavior_duration_stats_df['STD for baseline'] = [
            self.baseline_data_std] + [np.nan] * (len(behavior_duration_stats_df) - 1)

        return z_scored_data, z_scored_time, behavior_duration_stats_df

    def extract_behaviour_results(self, behaviours_to_export, params, z_scored_data=None):
        """
        Extract the results for the selected behaviours.

        Parameters:
        - behaviours_to_export: Set of behaviours to export.
        - params: Dictionary containing the parameters for each behaviour.
        - z_scored_data (Optional): The z-scored data for the selected column.

        Returns:
        - behaviours_results: Dictionary containing the extracted data for each behaviour.
        - time_ranges: Dictionary containing the time ranges for each behaviour.
        """
        behaviours_results = {}  # To hold the extracted data
        time_ranges = {}

        for behaviour_name in behaviours_to_export:
            behaviour_instances = params[behaviour_name]
            for behaviour_instance in behaviour_instances:
                pre_behaviour_time = float(
                    behaviour_instance['pre_behaviour_time'])
                post_behaviour_time = float(
                    behaviour_instance['post_behaviour_time'])
                behaviour_start_time = float(
                    behaviour_instance['behaviour_start_time'])

                if behaviour_name in behaviours_to_export and behaviour_name not in behaviours_results:
                    # Only create a new list if it doesn't exist yet
                    behaviours_results[behaviour_name] = []

                if behaviour_name not in behaviours_to_export:
                    continue

                if pre_behaviour_time:
                    start_time = behaviour_start_time - pre_behaviour_time
                    end_time = behaviour_start_time + post_behaviour_time

                    start_time_min = start_time / 60
                    end_time_min = end_time / 60
                    behaviour_start_time_min = behaviour_start_time / 60

                    num_samples = int(
                        (pre_behaviour_time + post_behaviour_time) * 10)
                    time_range = pd.Series(np.arange(-pre_behaviour_time, post_behaviour_time,
                                           (pre_behaviour_time + post_behaviour_time) / num_samples))

                    time_range[np.isclose(time_range, 0)] = 0

                    time_ranges[behaviour_name] = time_range

                    # Call extract_data function
                    if self.checkbox_state and z_scored_data is not None:
                        start_data, end_data = self.extract_data(start_time_min, behaviour_start_time_min, end_time_min,
                                                                 column='baselined_z_score', z_scored_data=z_scored_data)
                    else:
                        start_data, end_data = self.extract_data(start_time_min, behaviour_start_time_min, end_time_min,
                                                                 column=self.selected_column_var.get())

                    # Append the extracted data to behaviours_results
                    behaviours_results[behaviour_name].append(
                        (start_data, end_data))

        return behaviours_results, time_ranges

    def create_df_for_behaviours(self, behaviours_results, sorted_behaviours, time_ranges, folder_path):
        """
        Create a DataFrame for each behaviour and save it to a CSV file.

        Parameters:
        - behaviours_results: Dictionary containing the extracted data for each behaviour.
        - time_ranges: Dictionary containing the time ranges for each behaviour.
        - folder_path: Path to the folder where the CSV files will be saved.

        Returns:
        - df_list: List of DataFrames and their names for combining into a single CSV file.
        """

        df_list = []

        for behaviour_name in sorted_behaviours:
            data_list = behaviours_results[behaviour_name]
            df_combined = pd.DataFrame()  # Create a new DataFrame for each behaviour

            for instance_index, data in enumerate(data_list, start=1):
                start_data, end_data = data

                combined_data = pd.concat([start_data.reset_index(
                    drop=True), end_data.reset_index(drop=True)], axis=0)
                combined_data = combined_data.reset_index(drop=True)

                column_name = f"{behaviour_name}_{instance_index}"
                df_combined[column_name] = combined_data

            # Only add mean and sem columns if there is more than one instance of the behaviour
            if len(df_combined.columns) > 1:
                # Calculate the mean for each row
                df_combined['Mean'] = df_combined.mean(axis=1)

                # Calculate the SEM for each row
                df_combined['SEM'] = df_combined.sem(axis=1)

            time_range = time_ranges[behaviour_name]
            df_combined.insert(0, "Time (s)", time_range[:len(
                df_combined)].reset_index(drop=True))

            for col in df_combined.columns:
                df_combined[col] = df_combined[col].reindex(
                    time_range.index).reset_index(drop=True)

            behaviour_name = self.truncate_sheet_title(behaviour_name)
            behaviour_name = re.sub(r'[\\/*?:"<>|]', '_', behaviour_name)

            if self.export_options_container.combine_csv_var.get() == 1:

                # If combine_csv_var is 1, add the dataframe and its name to df_list
                df_list.append((df_combined, behaviour_name))

            else:
                # Add the selected column name to the file name
                selected_column_name = self.selected_column_var.get()
                results_file_name = f"{behaviour_name}_{
                    selected_column_name}_raw.csv"

                # Add '_baseline' if self.use_baseline_var is 1
                if self.checkbox_state:
                    results_file_name = results_file_name.replace(
                        "_raw.csv", "_baseline_raw.csv")

                results_file_path = os.path.join(
                    folder_path, results_file_name)
                df_combined.to_csv(results_file_path, index=False)

        return df_list

    def retrieve_static_params(self, params, behaviour_name):
        """
        Retrieve the static parameters for a given behaviour.

        Parameters:
        - params: Dictionary containing the parameters for each behaviour.
        - behaviour_name: The name of the behaviour.

        Returns:
        - pre_behaviour_time: The pre-behaviour time.
        - post_behaviour_time: The post-behaviour time.
        - bin_size: The bin size.
        """
        for behaviour_instance in params[behaviour_name]:
            pre_behaviour_time = int(
                behaviour_instance.get('pre_behaviour_time'))
            post_behaviour_time = int(
                behaviour_instance.get('post_behaviour_time'))
            bin_size = int(behaviour_instance.get('bin_size'))

            if bin_size is None:
                continue

        return pre_behaviour_time, post_behaviour_time, bin_size

    def get_metric_functions(self):
        """
        Calculate the metrics functions for the selected options.

        Returns:
        - metrics_functions: Dictionary containing the metrics functions.
        """
        metrics_functions = {}

        if self.export_options_container.use_auc_var.get():
            metrics_functions['auc'] = self.calculate_auc
        if self.export_options_container.use_max_amp_var.get():
            metrics_functions['max_amp'] = self.calculate_max_amp
        if self.export_options_container.use_mean_dff_var.get():
            metrics_functions['mean_dff'] = self.calculate_mean_dff
        return metrics_functions

    def calculate_metrics_for_bins(self, instances_data, expected_num_bins, metric_name, metric_func):
        """
        Calculate the metrics for each bin.

        Parameters:
        - instances_data: List of instances data.
        - expected_num_bins: Expected number of bins.
        - metrics_functions: Dictionary containing the metrics functions.

        Returns:
        - metric_values_per_bin: List of metric values for each bin.

        """
        metric_values_per_bin = []

        # Iterate over each bin
        for bin_idx in range(expected_num_bins):
            # Gather data for the current bin from all instances
            bin_data_across_instances = [
                instance[bin_idx] for instance in instances_data]

            # Calculate metric for each instance for the current bin
            instance_metric_values = [metric_func(
                instance_data) for instance_data in bin_data_across_instances]

            # If the metric is AUC and the result is somehow wrapped in a list or array, extract the scalar value
            if metric_name == 'auc' and isinstance(instance_metric_values[0], (list, np.ndarray)):
                instance_metric_values = [value[0]
                                          for value in instance_metric_values]

            # Compute the average metric value across all instances for this bin
            avg_value = sum(
                instance_metric_values) / len(instance_metric_values) if instance_metric_values else 0

            metric_values_per_bin.append(avg_value)

        return metric_values_per_bin

    def append_to_summary_data(self, summary_data, row_data):
        """
        Append a row of data to the summary_data.

        Parameters:
        - summary_data: List to hold the summary data.
        - row_data: List of data to append as a new row.
        """
        summary_data.append(row_data)

    def export_binned_data(self, sorted_behaviours, behaviours_results, params, file_path, folder_path, df_list, behavior_duration_stats_df):
        """
        Export the binned data to a CSV file.

        Parameters:
        - sorted_behaviours: List of behaviours sorted in alphabetical order.
        - behaviours_results: Dictionary containing the extracted data for each behaviour.
        - params: Dictionary containing the parameters for each behaviour.
        - file_path: Path to the photometry data file.
        - folder_path: Path to the folder where the CSV files will be saved.
        - df_list: List of DataFrames and their names for combining into a single CSV file.
        - df_duration: DataFrame containing duration data.
        """
        df_summary = self.generate_summary_data(
            sorted_behaviours, behaviours_results, params)

        df_list.insert(0, (behavior_duration_stats_df, "Event Duration"))

        if self.export_options_container.combine_csv_var.get() == 1:
            self.export_combined_csv(
                df_summary, df_list, file_path, folder_path, behaviours_results)
        else:
            self.export_separate_csv(df_summary, file_path, folder_path)

    def generate_summary_data(self, sorted_behaviours, behaviours_results, params):
        """
        Generate the summary data for the binned data.

        Parameters:
        - sorted_behaviours: List of behaviours sorted in alphabetical order.
        - behaviours_results: Dictionary containing the extracted data for each behaviour.
        - params: Dictionary containing the parameters for each behaviour.

        Returns:
        - df_summary: DataFrame containing the summary data.
        """
        summary_rows = []

        for behaviour_name in sorted_behaviours:
            data_list = behaviours_results[behaviour_name]

            # Retrieve parameters for the behaviour
            pre_behaviour_time, post_behaviour_time, bin_size = self.retrieve_static_params(
                params, behaviour_name)

            # Process and bin the data for the behaviour
            bin_labels, expected_num_bins, behaviour_instances_data = self.process_and_bin_data(
                data_list, pre_behaviour_time, post_behaviour_time, bin_size)

            # Add the behaviour name and bin labels to the summary
            summary_rows.append([behaviour_name] + bin_labels)

            # Segment the instances data based on the expected number of bins
            instances_data = [behaviour_instances_data[i:i + expected_num_bins]
                              for i in range(0, len(behaviour_instances_data), expected_num_bins)]

            # Get metric functions to apply to the data
            metrics_functions = self.get_metric_functions()

            # Calculate metrics for each bin and append to the summary
            for metric_name, metric_func in metrics_functions.items():
                metric_values_per_bin = self.calculate_metrics_for_bins(
                    instances_data, expected_num_bins, metric_name, metric_func)
                summary_rows.append([metric_name] + metric_values_per_bin)

            # Add a separator row (optional) for readability
            summary_rows.append([''] * (1 + len(bin_labels)))

        # Create a DataFrame from the collected summary rows
        df_summary = pd.DataFrame(summary_rows)

        return df_summary

    def prepare_combined_data(self, df_summary):
        """
        Prepare data for combined CSV export. Add a separator row (NaN) after the summary data.

        Parameters:
        - df_summary: DataFrame containing the summary data for all behaviors.

        Returns:
        - summary_data: List containing the prepared DataFrame and a separator row.
        """
        # Add the summary data directly
        summary_data = [df_summary]

        # Append a row of NaNs as a separator
        separator_row = pd.DataFrame(
            np.nan, columns=df_summary.columns, index=[0])
        summary_data.append(separator_row)

        return summary_data

    def export_combined_csv(self, df_summary, df_list, file_path, folder_path, behaviours_results):
        """
        Export combined CSV for all behaviors, including summary data.

        Parameters:
        - df_summary: DataFrame containing the summary data for all behaviors.
        - df_list: List of additional DataFrames to be exported.
        - file_path: Path to the original file.
        - folder_path: Path to the folder where the CSV files will be saved.
        - behaviours_results: Dictionary of extracted data for each behavior.
        """
        summary_data = []

        combined_data = self.prepare_combined_data(df_summary)
        summary_data.extend(combined_data)

        # Concatenate the collected data into a single DataFrame
        df_summary_combined = pd.concat(summary_data, ignore_index=True)

        # Ensure the columns are anonymous (i.e., remove column headers)
        df_summary_combined.columns = [''] * df_summary_combined.shape[1]

        # Reset index and clear column names
        df_summary_combined = df_summary_combined.reset_index(drop=True)
        df_summary_combined.columns.name = None

        # Add the summary results at the start of the DataFrame list
        df_list.insert(0, (df_summary_combined, "Summary Results"))

        # Generate the output file name and save the file
        output_file_name = self.get_output_file_name(file_path, folder_path)
        self.save_to_excel(df_list, output_file_name)

        # Optionally format the Excel file
        self.format_excel(output_file_name, behaviours_results)

    def get_output_file_name(self, file_path, folder_path):
        original_file_name = os.path.splitext(os.path.basename(file_path))[0]
        selected_column_name = self.selected_column_var.get()
        original_file_name += f"_{selected_column_name}"

        if self.checkbox_state:
            original_file_name += f"_baseline_{
                self.data_selection_frame.baseline_start_entry.get()}"

        return os.path.join(folder_path, f"{original_file_name}.xlsx")

    def save_to_excel(self, df_list, output_file_name):
        with pd.ExcelWriter(output_file_name, engine='openpyxl') as writer:
            for df, sheet_name in df_list:
                sheet_name = sheet_name.replace("/", "_")
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def format_excel(self, output_file_name, behaviours_results):
        wb = load_workbook(output_file_name)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name == 'Event Duration':
                self.bold_first_row(ws)
                self.bold_first_column(ws)
            else:
                if sheet_name == wb.sheetnames[0]:
                    ws.delete_rows(1)
                self.bold_first_column(ws)
                self.bold_behavior_name_rows(ws, behaviours_results)
            self.remove_borders(ws)
        wb.save(output_file_name)

    def export_separate_csv(self, df_summary, file_path, folder_path):
        file_name = os.path.basename(file_path)
        file_name_without_extension = os.path.splitext(file_name)[0]
        selected_column_name = self.selected_column_var.get()
        file_name_without_extension += f"_{selected_column_name}"

        if self.checkbox_state:
            file_name_without_extension += f"_baseline_{
                self.data_selection_frame.baseline_start_entry.get()}"

        if self.export_options_container.use_binned_data_var.get() == 1:
            file_name_without_extension += "_binned"

        df_results_path = os.path.join(
            folder_path, f"{file_name_without_extension}_summary.csv")

        df_summary.to_csv(df_results_path, index=False, header=False)

        # TODO REMOVE THIS?

    def bold_first_row(self, ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    def bold_first_column(self, ws):
        for cell in ws['A']:
            cell.font = Font(bold=True)

    def bold_behavior_name_rows(self, ws, behaviours_results):
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
            if row[0].value in behaviours_results:
                for cell in row:
                    cell.font = Font(bold=True)

    def remove_borders(self, ws):
        for cell in ws[1]:
            cell.border = Border(left=Side(style='none'),
                                 right=Side(style='none'),
                                 top=Side(style='none'),
                                 bottom=Side(style='none'))

    def process_and_bin_data(self, data_list, pre_behaviour_time, post_behaviour_time, bin_size):
        """
        Process and bin the data for a given behaviour.

        Parameters:
        - summary_data: List to hold the summary data.
        - behaviour_name: The name of the behaviour.
        - data_list: List of data for the behaviour.
        - pre_behaviour_time: The pre-behaviour time.
        - post_behaviour_time: The post-behaviour time.
        - bin_size: The bin size.
        - binned_data_list: List to hold the binned data.

        Returns:
        - bin_labels: List of bin labels.
        - metrics_functions: Dictionary containing the metrics functions to use.
        - expected_num_bins: The expected number of bins.
        - instances_data: List of instances data.
        - behaviour_instances_data: List of binned data for each behaviour instance.
        """

        # Initialize a list to hold the binned data
        behaviour_instances_data = []

        for start_data, end_data in data_list:
            combined_data_array = np.concatenate(
                [start_data.values, end_data.values])
            sampling_rate = len(
                combined_data_array) // (pre_behaviour_time + post_behaviour_time)
            num_bins = len(
                combined_data_array) // (bin_size * sampling_rate)

            for bin_idx in range(num_bins):
                start_idx = bin_idx * bin_size * sampling_rate
                end_idx = start_idx + bin_size * sampling_rate
                binned_data = combined_data_array[start_idx:end_idx].reshape(
                    -1, bin_size * sampling_rate)
                behaviour_instances_data.append(binned_data)

            # Bin Labels
            num_bins_total = (pre_behaviour_time +
                              post_behaviour_time) // bin_size
            bin_ranges = np.linspace(-pre_behaviour_time,
                                     post_behaviour_time, num_bins_total, endpoint=False)
            bin_labels = sorted(bin_ranges, key=lambda x: float(x))
            bin_labels = [
                f'{start} - {start + bin_size}' for start in bin_labels]

        return bin_labels, num_bins_total, behaviour_instances_data

    def extract_data_from_photometry(self, file_path, params):
        """
        Extract data from the photometry data file.

        Parameters:
        - file_path: Path to the photometry data file.
        - params: Parameters for data extraction and processing.
        """

        # Create a folder to store the extracted data
        folder_path = self.create_extraction_folder(file_path)

        behaviours_metrics = {}  # To hold the calculated metrics
        bin_size = params.get("bin_size")
        behaviours_to_export = params.get("behaviours_to_export", [])

        behavior_duration_stats_df = self.extract_duration_data()

        z_scored_data, z_scored_time = None, None

        if self.checkbox_state:
            z_scored_data, z_scored_time, behavior_duration_stats_df = self.handle_zscore_checkbox_export(
                behavior_duration_stats_df)

        behaviours_results, time_ranges = self.extract_behaviour_results(
            behaviours_to_export, params, z_scored_data)

        sorted_behaviours = sorted(behaviours_results.keys())

        df_list = self.create_df_for_behaviours(
            behaviours_results, sorted_behaviours, time_ranges, folder_path)

        if self.export_options_container.use_binned_data_var.get() == 1:
            self.export_binned_data(sorted_behaviours, behaviours_results,
                                    params, file_path, folder_path, df_list, behavior_duration_stats_df)

        else:
            temp_data = []
            for behaviour_name in behaviours_results:
                # Iterating over the list of dictionaries
                for behaviour_instance in params[behaviour_name]:
                    pre_behaviour_time = float(
                        behaviour_instance['pre_behaviour_time'])
                    post_behaviour_time = float(
                        behaviour_instance['post_behaviour_time'])

                    for start_data, end_data in behaviours_results[behaviour_name]:
                        # Create a dictionary with the always-included fields
                        data_dict = {
                            "behaviour_name": behaviour_name,
                            "pre_behaviour_time": pre_behaviour_time,
                            "post_behaviour_time": post_behaviour_time,
                            "bin_size": bin_size,
                        }

                        # Conditionally include calculations based on whether they should be performed
                        if self.export_options_container.use_auc_var.get():
                            data_dict["start_auc"] = self.calculate_auc(
                                start_data)
                            data_dict["end_auc"] = self.calculate_auc(end_data)

                        if self.export_options_container.use_max_amp_var.get():
                            data_dict["start_max_amp"] = self.calculate_max_amp(
                                start_data)
                            data_dict["end_max_amp"] = self.calculate_max_amp(
                                end_data)

                        if self.export_options_container.use_mean_dff_var.get():
                            data_dict["start_mean_dff"] = self.calculate_mean_dff(
                                start_data)
                            data_dict["end_mean_dff"] = self.calculate_mean_dff(
                                end_data)

                        # Append the dictionary to temp_data
                        temp_data.append(data_dict)

            for data in temp_data:
                behaviour_name = data["behaviour_name"]

                if behaviour_name not in behaviours_metrics:
                    behaviours_metrics[behaviour_name] = {}

                for key in data.keys():  # Loop through keys in the data dictionary
                    if key not in ["behaviour_name", "pre_behaviour_time", "post_behaviour_time", "bin_size"]:
                        if key not in behaviours_metrics[behaviour_name]:
                            behaviours_metrics[behaviour_name][key] = []
                        behaviours_metrics[behaviour_name][key].append(
                            data[key])

                for key in ["pre_behaviour_time", "post_behaviour_time", "bin_size"]:
                    behaviours_metrics[behaviour_name][key] = data[key]

            # Create an empty list to store the individual dataframes for each metric
            dfs = []

            def add_empty_column(df):
                """
                Add an empty column to the DataFrame.

                Parameters:
                - df: The DataFrame to which the empty column will be added.

                Returns:
                - df: The DataFrame with the added empty column.
                """
                df[''] = ""
                return df

            # Create DataFrame for times and bin size
            df_times = pd.DataFrame(
                [(k, v["pre_behaviour_time"], v["post_behaviour_time"])
                 for k, v in behaviours_metrics.items()],
                columns=['Behaviour', 'Pre Behaviour Time', 'Post Behaviour Time'])

            dfs.append(add_empty_column(df_times))

            # Conditionally create dataframes for the calculated metrics
            if 'start_auc' in behaviours_metrics[next(iter(behaviours_metrics))]:
                df_auc = pd.DataFrame([(k, np.mean(v.get('start_auc', [np.nan])), np.mean(v.get('end_auc', [np.nan]))) for k, v in behaviours_metrics.items()],
                                      columns=['Behaviour', 'Start AUC', 'End AUC'])
                dfs.append(add_empty_column(df_auc))

            if 'start_max_amp' in behaviours_metrics[next(iter(behaviours_metrics))]:
                df_max_amp = pd.DataFrame(
                    [(k, np.mean(v.get('start_max_amp', [np.nan])), np.mean(
                        v.get('end_max_amp', [np.nan]))) for k, v in behaviours_metrics.items()],
                    columns=['Behaviour', 'Start Max AMP', 'End Max AMP'])
                dfs.append(add_empty_column(df_max_amp))

            if 'start_mean_dff' in behaviours_metrics[next(iter(behaviours_metrics))]:
                df_mean_dff = pd.DataFrame(
                    [(k, np.mean(v.get('start_mean_dff', [np.nan])), np.mean(
                        v.get('end_mean_dff', [np.nan]))) for k, v in behaviours_metrics.items()],
                    columns=['Behaviour', 'Start Mean dF/F', 'End Mean dF/F'])
                dfs.append(add_empty_column(df_mean_dff))

            # Concatenate the dataframes to create the final df_results
            df_results = pd.concat(dfs, axis=1)

            if self.checkbox_state:
                baseline_start_time = float(
                    self.data_selection_frame.baseline_start_entry.get())
                baseline_end_time = float(
                    self.data_selection_frame.baseline_end_entry.get())
                df_results["Baseline Start Time"] = ""
                df_results["Baseline End Time"] = ""
                df_results.at[0, "Baseline Start Time"] = baseline_start_time
                df_results.at[0, "Baseline End Time"] = baseline_end_time

            # Assuming df_list is defined somewhere above in your code
            # Add df_results as the first item in df_list
            df_list.insert(0, (df_results, 'Summary Results'))

            if self.export_options_container.combine_csv_var.get() == 1:
                # Get the original filename without the extension
                original_file_name = os.path.splitext(
                    os.path.basename(file_path))[0]

                # Add the selected column name to the original filename
                selected_column_name = self.selected_column_var.get()
                original_file_name += f"_{selected_column_name}"

                # Add '_baseline' if self.use_baseline_var is 1
                if self.checkbox_state:
                    original_file_name += f"_baseline_{
                        self.data_selection_frame.baseline_start_entry.get()}"

                output_file_name = self.get_output_file_name(
                    file_path, folder_path)
                # If combine_csv_var is 1, write all dataframes from df_list to the same Excel file
                with pd.ExcelWriter(os.path.join(folder_path, output_file_name), engine='openpyxl') as writer:
                    for df, sheet_name in df_list:
                        sheet_name = sheet_name.replace("/", "_")
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Load the workbook
                wb = load_workbook(os.path.join(folder_path, output_file_name))

                # Loop through all sheets
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Clear border from column headers
                    for cell in ws[1]:
                        cell.border = Border(left=Side(style='none'),
                                             right=Side(style='none'),
                                             top=Side(style='none'),
                                             bottom=Side(style='none'))

                # Save the workbook
                wb.save(os.path.join(folder_path, output_file_name))

            else:
                file_name = os.path.basename(file_path)
                file_name_without_extension = os.path.splitext(
                    file_name)[0]  # Extract file name without extension

                # Add the selected column name to the file name
                selected_column_name = self.selected_column_var.get()
                file_name_without_extension += f"_{selected_column_name}"

                # Add '_baseline' if self.use_baseline_var is 1
                if self.checkbox_state:
                    file_name_without_extension += f"_baseline_{
                        self.data_selection_frame.baseline_start_entry.get()}"

                # Create the summary results file path
                df_results_path = os.path.join(
                    folder_path, f"{file_name_without_extension}_summary.csv")
                df_results.to_csv(df_results_path, index=False)

    def extract_data(self, start_time_min, behaviour_time_min, end_time_min, column, z_scored_data=None):
        """
        Extract data from specific time points in a CSV file and return the data as a pandas DataFrame.

        Parameters:
        - start_time_min: Start time in minutes.
        - behaviour_time_min: Behaviour time in minutes.
        - end_time_min: End time in minutes.
        - column: The column from which to extract data.
        - z_scored_data: Optional z-scored data.

        Returns:
        - start_data: Data from the start to the behaviour time.
        - end_data: Data from the behaviour time to the end time.
        """
        df = self.dataframe

        if column == 'baselined_z_score' and z_scored_data is not None:
            df['baselined_z_score'] = z_scored_data
            time_column = df['z_scored_time']
        else:
            # Use the first column for the time values, regardless of its name
            time_column = df.iloc[:, 0]

        # Now, when you need to find the index of the nearest time value, use the first column
        start_time_idx = (time_column - start_time_min).abs().idxmin()
        behaviour_time_idx = (time_column - behaviour_time_min).abs().idxmin()
        end_time_idx = (time_column - end_time_min).abs().idxmin()

        start_data = df.loc[start_time_idx:behaviour_time_idx,
                            column].reset_index(drop=True)
        end_data = df.loc[behaviour_time_idx:end_time_idx,
                          column].reset_index(drop=True)

        return start_data, end_data

    def check_and_prepare_parameters(self, table_key):
        """
        Check if the parameters are valid and prepare the parameters for the data extraction process.

        Parameters:
        - table_key: The key for the table containing the parameters.

        Returns:
        - params: A dictionary of parameters for data extraction, or None if there are issues with the parameters.

        Raises:
        - ValueError: If there are issues with the parameters.
        """
        if table_key not in self.tables:
            return

        df = self.tables[table_key]

        # Filter out rows where the Start Time is negative
        df['Start Time'] = df['Start Time'].astype(float)
        df = df[df['Start Time'] >= 0]

        params = {}
        behaviours_to_export = set()
        params_to_extract = []
        bin_size = None
        missing_pre_behaviour = set()
        missing_post_behaviour = set()
        # Store behaviours that failed the total time divisibility check
        not_divisible_behaviours = set()

        if self.checkbox_state:
            baseline_start_time_min = float(
                self.data_selection_frame.baseline_start_entry.get()) / 60
            self.z_scored_data, self.z_scored_time = self.calculate_z_score()
            params["baseline_start_time_min"] = baseline_start_time_min

        for _, row in df.iterrows():
            behaviour_name = row['Behaviour Name']
            status_var = self.behaviour_display_status.get(behaviour_name)

            if status_var and status_var.get() == 1:  # If the checkbox is ticked
                pre_behaviour_time = row['Pre Behaviour Time']
                post_behaviour_time = row['Post Behaviour Time']
                bin_size = row['Bin Size']

                # Create a dictionary for this instance of the behavior
                this_behaviour_instance = {
                    "pre_behaviour_time": pre_behaviour_time,
                    "post_behaviour_time": post_behaviour_time,
                    "bin_size": bin_size,
                    "behaviour_start_time": float(row['Start Time'])
                }

                # Add this instance to the list of instances for this behavior
                if behaviour_name not in params:
                    params[behaviour_name] = []
                params[behaviour_name].append(this_behaviour_instance)

                if pre_behaviour_time == "":
                    missing_pre_behaviour.add(behaviour_name)
                elif post_behaviour_time == "":
                    missing_post_behaviour.add(behaviour_name)
                else:
                    behaviour_start_time = row['Start Time']

                    if not self.check_total_time_divisible(float(pre_behaviour_time), float(post_behaviour_time),
                                                           bin_size):
                        not_divisible_behaviours.add(behaviour_name)
                    else:
                        behaviours_to_export.add(behaviour_name)

        if missing_pre_behaviour:
            missing_behaviours_str = ", ".join(missing_pre_behaviour)
            messagebox.showwarning(
                "Pre-behaviour Time Missing", f"Pre-behaviour time is missing for the following behaviours: {missing_behaviours_str}")
            return None

        if missing_post_behaviour:
            missing_post_behaviours_str = ", ".join(missing_post_behaviour)
            messagebox.showwarning("Post-behaviour Time Missing",
                                   f"Post-behaviour time is missing for behaviour: {missing_post_behaviours_str}")
            return None

        if not_divisible_behaviours:
            not_divisible_behaviours_str = ", ".join(not_divisible_behaviours)
            messagebox.showwarning("Total Time Not Divisible",
                                   f"Total time for the following behaviours is not divisible by the bin size: {not_divisible_behaviours_str}")
            return None

        params["behaviours_to_export"] = behaviours_to_export
        params["params_to_extract"] = params_to_extract

        return params

    def check_total_time_divisible(self, pre_behaviour_time, post_behaviour_time, bin_size):
        """
        Check if the total time is divisible by the bin size.

        Parameters:
        - pre_behaviour_time: The pre-behaviour time.
        - post_behaviour_time: The post-behaviour time.
        - bin_size: The bin size.

        Returns:
        - is_divisible: Boolean indicating whether the total time is divisible by the bin size.
        """
        total_time = int(pre_behaviour_time + post_behaviour_time)
        is_divisible = total_time % int(bin_size) == 0

        return is_divisible

    # class GraphDisplay:

    def configure_canvas(self, event):
        """
        Configure the canvas to allow scrolling.

        Parameters:
        - event: The event that triggered the configuration.
        """
        canvas_height = min(
            event.height, self.settings_container_frame.winfo_height())
        self.canvas.configure(scrollregion=self.canvas.bbox(
            "all"), height=canvas_height)

    def refresh_graph_display(self):
        """Refresh the graph display"""
        self.handle_figure_display_selection(None)

    def get_time_scale(self, time_unit):
        """
        Get the time scale factor based on the time unit.

        Parameters:
        - time_unit: The unit of time ('minutes', 'seconds', 'hours').

        Returns:
        - scale_factor: The scale factor for the time unit.
        """
        if time_unit == 'minutes':
            return 1
        elif time_unit == 'seconds':
            return 60
        elif time_unit == 'hours':
            return 1 / 60
        else:
            return 1

    def convert_and_retrieve_time(self, time_data, return_label=False):
        """
        Convert time data based on the selected time unit and optionally return the x-label.

        Parameters:
        - time_data: The time data to convert.
        - return_label: Boolean flag indicating whether to return the x-label.

        Returns:
        - converted_time_data: The converted time data.
        - label (optional): The x-label for the time unit.
        """
        time_unit = self.graph_settings_container_instance.time_unit_menu.get()

        conversion_factors = {
            'minutes': (1, 'Time (min)'),
            'seconds': (60, 'Time (s)'),
            'hours': (1 / 60, 'Time (h)')
        }

        factor, label = conversion_factors.get(time_unit, (1, 'Time (min)'))

        converted_time_data = time_data * self.get_time_scale(time_unit)

        if return_label:
            return converted_time_data, label
        else:
            return converted_time_data

    def determines_ax_tick_spacing(self, ax):
        """
        Determine the tick spacing for the x and y axes.

        Parameters:
        - ax: The axis on which to set the tick spacing.
        """
        x_ticks_str = self.graph_settings_container_instance.x_gridlines_var.get()
        y_ticks_str = self.graph_settings_container_instance.y_gridlines_var.get()

        # Only adjust if the user has specified a value
        if x_ticks_str and x_ticks_str.strip():
            x_ticks = float(x_ticks_str)
            if x_ticks > 0:
                ax.xaxis.set_major_locator(MultipleLocator(x_ticks))
            else:
                print("X Ticks must be greater than 0")
        else:
            # Let the plot determine the spacing
            ax.xaxis.set_major_locator(AutoLocator())

        if y_ticks_str and y_ticks_str.strip():
            y_ticks = float(y_ticks_str)
            if y_ticks > 0:
                ax.yaxis.set_major_locator(MultipleLocator(y_ticks))
            else:
                print("Y Ticks must be greater than 0")
        else:
            # Let the plot determine the spacing
            ax.yaxis.set_major_locator(AutoLocator())

    def retrieve_and_process_behaviour_data(self, current_df=None):
        """
        Retrieve and process behaviour data from a pandas DataFrame.

        Parameters:
        - current_df: Optional DataFrame to process. If None, use the current table DataFrame.

        Returns:
        - behaviours: List of behaviour names.
        - start_times: List of start times.
        - end_times: List of end times.
        - start_times_min: List of start times in minutes.
        - end_times_min: List of end times in minutes.

        Raises:
        - ValueError: If the start or end time is not a valid
        """
        if current_df is None:
            # Retrieve data from pandas DataFrame if available
            all_items = self.tables.get(
                self.current_table_key, pd.DataFrame()).to_dict('records')
        else:
            all_items = current_df.to_dict('records')

        # Extract data using list comprehensions
        behaviours = [record['Behaviour Name'] for record in all_items]
        start_times = [record['Start Time'] for record in all_items]
        end_times = [record['End Time'] for record in all_items]

        # Convert start times to minutes
        start_times_min = [float(time) / 60 for time in start_times]

        # Convert end times to minutes with error handling
        end_times_min = []
        for time in end_times:
            try:
                end_times_min.append(float(time) / 60)
            except (ValueError, TypeError):
                end_times_min.append(None)

        return behaviours, start_times, end_times, start_times_min, end_times_min

    def handle_figure_display_selection(self, event):
        """
        Handle the figure display selection.

        Parameters:
        - event: The event that triggered the selection.
        """
        selected_option = self.figure_display_dropdown.get()
        matplotlib.pyplot.close()

        # get user's time unit selection
        time_unit = self.graph_settings_container_instance.time_unit_menu.get()

        if not self.is_file_parsed and self.data_selection_frame.baseline_button_pressed:
            print("Please parse the file first.")
            self.data_already_adjusted = True

        if hasattr(self, "figure_canvas") and self.figure_canvas is not None:
            self.figure_canvas.get_tk_widget().pack_forget()
            self.figure_canvas.get_tk_widget().destroy()
            self.toolbar.destroy()

        # Create the figure and axes
        self.fig, ax = plt.subplots(figsize=(6, 4))

        self.ax = ax

        # Hide the top and right axis lines
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        if selected_option == "Z-scored data" and self.checkbox_state:
            self.plot_z_scored_data(ax)

        elif selected_option == "Full Trace Display" or (selected_option == "Behaviour Mean and SEM" and self.behaviour_choice_graph.get() == ""):
            self.plot_full_trace(ax)
            self.default_y_limits = ax.get_ylim()
            self.original_xticks = self.ax.get_xticks().tolist()
            self.original_xticklabels = [
                label.get_text() for label in self.ax.get_xticklabels()]
            if selected_option == "Behaviour Mean and SEM":
                tk.messagebox.showinfo(
                    "No Behaviour Selected", "Please select a behaviour to display the mean and SEM.")
                self.figure_display_dropdown.set("Full Trace Display")

        elif selected_option == "Single Row Display":
            self.plot_single_row(ax)
            ax.set_ylim(auto=True)

        elif selected_option == "Behaviour Mean and SEM":
            self.plot_mean_and_sem_trace(ax)
            scale_factor = self.convert_and_retrieve_time(1)
            ax.set_xlim(self.min_x_for_behaviour_mean_and_sem * scale_factor,
                        self.max_x_for_behaviour_mean_and_sem * scale_factor)

        # Determine tick spacing for the x and y axes after plotting
        self.determines_ax_tick_spacing(ax)

        if selected_option not in ["Single Row Display", "Behaviour Mean and SEM"]:
            current_xlim = ax.get_xlim()  # get current x-axis limits
            # set x-axis limits to start at 0
            ax.set_xlim(left=current_xlim[0], right=current_xlim[1])

        # Assuming self.fig is your matplotlib figure object
        fig_dpi = self.fig.dpi
        fig_width, fig_height = self.fig.get_size_inches()
        canvas_width, canvas_height = fig_width * fig_dpi, fig_height * fig_dpi

        # Convert the matplotlib figure to a Tkinter-compatible object
        self.figure_canvas = FigureCanvasTkAgg(
            self.fig, master=self.graph_canvas)
        # Configure the figure canvas size explicitly
        self.figure_canvas.get_tk_widget().config(
            width=canvas_width, height=canvas_height)
        self.figure_canvas.draw()

        # Place the figure canvas in the graph_canvas
        self.figure_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Create the toolbar
        self.toolbar = NavigationToolbar2Tk(
            self.figure_canvas, self.graph_canvas)
        self.toolbar.update()

        # Customize the toolbar background
        self.toolbar.config(background="snow")
        self.toolbar._message_label.config(background="snow")

        # Customize the x and y coordinate labels
        self.toolbar._message_label.config(
            foreground="black", font=("Arial", 10))

        # Remove the canvas border
        self.figure_canvas.get_tk_widget().configure(
            borderwidth=0, highlightthickness=0)

        # Remove the toolbar border
        self.toolbar.configure(background="snow", bd=0)
        self.toolbar._message_label.configure(background="snow", bd=0)

        # Place the toolbar in the graph_canvas
        self.toolbar.pack(side="top", fill="x")
        self.settings_manager.save_variables()

    def plot_z_scored_data(self, ax):
        """
        Plot the z-scored data.

        Parameters:
        - ax: The axis on which to plot the data.
        """
        if not self.baseline_button_pressed:
            self.calculate_z_score()

        # Check for required columns in dataframe
        if 'baselined_z_score' not in self.dataframe or 'z_scored_time' not in self.dataframe:
            print("Required data columns are missing in the dataframe.")
            return

        # Retrieve z-scored data and its associated time directly from the dataframe
        z_scored_data = self.dataframe['baselined_z_score']
        z_scored_time = self.dataframe['z_scored_time']

        # Retrieve the x-axis label and converted time data based on user's time unit selection
        z_scored_time_copy = z_scored_time.copy()
        converted_time_data, x_label = self.convert_and_retrieve_time(
            z_scored_time_copy, return_label=True)
        ax.set_xlabel(x_label)

        behaviours = None
        adjusted_time = converted_time_data

        if self.original_table is not None and not self.original_table.empty:
            current_df = self.original_table.copy()
            self.update_table(current_df)

            behaviours, _, _, start_times_min, end_times_min = self.retrieve_and_process_behaviour_data(
                current_df)

            # Initialize or check time attributes
            self.initialize_or_check_time_attributes(
                start_times_min, end_times_min)

            zeroing_active = self.graph_settings_container_instance.zero_x_axis_checkbox_var.get() == 1

            if zeroing_active:
                adjusted_start_times_min, adjusted_end_times_min, adjusted_time = self.handle_zeroing(
                    behaviours, start_times_min, end_times_min, converted_time_data
                )
            elif self.checkbox_state:
                adjusted_start_times_min = start_times_min
                adjusted_end_times_min = end_times_min
            else:
                adjusted_start_times_min = self.original_start_times_min if self.original_start_times_min else start_times_min
                adjusted_end_times_min = self.original_end_times_min if self.original_end_times_min else end_times_min

        # Plot the data
        ax.plot(adjusted_time, z_scored_data, color=self.settings_manager.selected_trace_color, linewidth=float(
            self.graph_settings_container_instance.line_width_entry.get()))
        ax.set_ylabel('Z-score')

        self.xlim_min = adjusted_time.min()
        self.xlim_max = adjusted_time.max()

        # Add transparent boxes if behaviors are present
        if behaviours:
            self.add_transparent_boxes(
                ax, z_scored_data, behaviours, adjusted_start_times_min, adjusted_end_times_min, adjusted_time)

        ax.set_xlim(self.xlim_min, self.xlim_max)

    def plot_full_trace(self, ax):
        """
        Plot the full trace.

        Parameters:
        - ax: The axis on which to plot the full trace.
        """
        selected_column = self.selected_column_var.get()
        # Create a copy to prevent modifying the original
        time = self.dataframe.iloc[:, 0].copy()
        data = self.dataframe[selected_column]

        behaviours, _, _, start_times_min, end_times_min = self.retrieve_and_process_behaviour_data()

        # Retrieve the x-axis label and converted time data based on user's time unit selection
        converted_time_data, x_label = self.convert_and_retrieve_time(
            time, return_label=True)
        ax.set_xlabel(x_label)

        # Initialize or check time attributes
        self.initialize_or_check_time_attributes(
            start_times_min, end_times_min)

        zeroing_active = self.graph_settings_container_instance.zero_x_axis_checkbox_var.get() == 1

        if zeroing_active and behaviours:
            adjusted_start_times_min, adjusted_end_times_min, adjusted_time = self.handle_zeroing(
                behaviours, start_times_min, end_times_min, converted_time_data
            )
        else:
            adjusted_start_times_min = self.original_start_times_min
            adjusted_end_times_min = self.original_end_times_min
            adjusted_time = converted_time_data

        ax.plot(adjusted_time, data, color=self.settings_manager.selected_trace_color, linewidth=float(
            self.graph_settings_container_instance.line_width_entry.get()))
        ax.set_ylabel(selected_column)

        self.xlim_min = adjusted_time.min()
        self.xlim_max = adjusted_time.max()

        if behaviours:
            self.add_transparent_boxes(
                ax, data, behaviours, adjusted_start_times_min, adjusted_end_times_min, adjusted_time)

        ax.set_xlim(self.xlim_min, self.xlim_max)

    def plot_single_row(self, ax):
        """
        Plot the selected row in the table.

        Parameters:
        - ax: The axis on which to plot the selected row.
        """
        start_point = self.start_time - self.pre_behaviour_time
        end_point = self.start_time + self.post_behaviour_time
        selected_data_to_plot = self.dataframe.loc[(
            self.dataframe.iloc[:, 0] >= start_point) & (self.dataframe.iloc[:, 0] <= end_point)]

        # if use_baseline_var is 1, use the z_score column. Otherwise, use the original column
        if self.checkbox_state:
            selected_column = self.calculate_z_score()
        else:
            selected_column = self.selected_column_var.get()

        # Retrieve the x-axis label and converted time data based on user's time unit selection
        time_data = selected_data_to_plot.iloc[:, 0].copy()
        converted_time_data, x_label = self.convert_and_retrieve_time(
            time_data, return_label=True)
        ax.set_xlabel(x_label)

        # Plot the single row
        onset_line_style = self.graph_settings_container_instance.onset_line_style_combobox.get()
        ax.plot(converted_time_data, selected_data_to_plot[selected_column], color=self.settings_manager.selected_trace_color, linewidth=float(
            self.graph_settings_container_instance.line_width_entry.get()))
        ax.set_ylabel(selected_column)
        ax.axvline(x=0, color=self.settings_manager.selected_line_color,
                   linestyle=onset_line_style)

        # Add transparent boxes to the graph
        start_times_min = [self.start_time]
        end_times_min = [self.end_time]
        item = self.table_treeview.selection()[0]  # Get selected item
        row_behaviour_name = self.table_treeview.item(item)['values'][2]
        self.add_transparent_boxes(ax, selected_data_to_plot[selected_column], [
                                   row_behaviour_name], start_times_min, end_times_min, start_point, end_point)

    def plot_mean_and_sem_trace(self, ax):
        """
        Plot the mean and SEM of the selected behaviour.

        Parameters:
        - ax: The axis on which to plot the mean and SEM trace.
        """
        behaviour_occurrences, column_used, pre_behaviour_times, post_behaviour_times = self.fetch_behaviour_data()

        # Get start and end times for use in the add_duration_box function
        start_times = [occurrence[0] for occurrence in behaviour_occurrences]
        end_times = [occurrence[1] for occurrence in behaviour_occurrences]

        start_time_adjusted = -pre_behaviour_times[0]
        end_time_adjusted = post_behaviour_times[0]

        # Set these as instance variables
        self.current_start_times = start_times
        self.current_end_times = end_times

        behaviour_data_by_instance, time_points, _, _ = self.process_behaviour_data(
            behaviour_occurrences, column_used)

        mean_sem_df = self.generate_behaviour_graph(
            ax, behaviour_data_by_instance, time_points, start_time_adjusted, end_time_adjusted)

        # Update the cache with the new mean_sem_df
        current_behavior = self.behaviour_choice_graph.get()
        if current_behavior in self.duration_data_cache:
            self.duration_data_cache[current_behavior]["mean_sem_df"] = mean_sem_df
        else:
            print(
                f"Warning: Behavior '{current_behavior}' not found in cache to update mean_sem_df.")

        # Add a box with the number of instances
        if self.graph_settings_container_instance.num_instances_box_var.get():
            num_instances = len(behaviour_data_by_instance)
            ax.text(0.95, 0.95, f'n = {num_instances}', transform=ax.transAxes, ha='right', va='top',
                    bbox=dict(facecolor='white', edgecolor='black', boxstyle='square,pad=0.5'))

        self.min_x_for_behaviour_mean_and_sem = start_time_adjusted
        self.max_x_for_behaviour_mean_and_sem = end_time_adjusted

        onset_line_style = self.graph_settings_container_instance.onset_line_style_combobox.get()
        ax.axvline(x=0, color=self.graph_settings_container_instance.selected_line_color, linestyle=onset_line_style,
                   linewidth=self.graph_settings_container_instance.onset_line_thickness_entry.get())

        if self.graph_settings_container_instance.limit_axis_range_var.get():
            y_min = self.graph_settings_container_instance.y_axis_min_var.get()
            y_max = self.graph_settings_container_instance.y_axis_max_var.get()
            ax.set_ylim(self.ax.set_ylim(float(y_min), float(y_max)))

        if hasattr(self, 'bar_items') and self.bar_items:
            self.update_duration_box(ax, mean_sem_df)
        else:
            self.add_duration_box(ax, mean_sem_df)

        # Adjust the layout to ensure that everything fits well
        plt.tight_layout()

    def initialize_or_check_time_attributes(self, start_times_min, end_times_min):
        """
        Initialize the original_start_times_min and original_end_times_min attributes if they don't exist.

        Parameters:
        - start_times_min: List of start times in minutes.
        - end_times_min: List of end times in minutes.
        """
        if not hasattr(self, 'original_start_times_min') or not self.original_start_times_min:
            self.original_start_times_min = start_times_min.copy()

        if not hasattr(self, 'original_end_times_min') or not self.original_end_times_min:
            self.original_end_times_min = end_times_min.copy()

    def save_and_close(self, popup=None, close=True):
        """
        Save the axis range variables and close the popup window.

        Parameters:
        - popup: The popup window to close (default is None).
        - close: Boolean indicating whether to close the popup window (default is True).
        """
        x_min = self.graph_settings_container_instance.x_axis_min_var.get()
        x_max = self.graph_settings_container_instance.x_axis_max_var.get()
        y_min = self.graph_settings_container_instance.y_axis_min_var.get()
        y_max = self.graph_settings_container_instance.y_axis_max_var.get()

        selected_option = None
        if hasattr(self, 'figure_display_dropdown'):
            selected_option = self.figure_display_dropdown.get()

        if self.graph_settings_container_instance.limit_axis_range_var.get():  # Check if the checkbox is ticked
            # Check for empty strings and set axis limits accordingly
            if x_min and x_max and selected_option != "Behaviour Mean and SEM":
                self.ax.set_xlim(float(x_min), float(x_max))
            if y_min and y_max:
                self.ax.set_ylim(float(y_min), float(y_max))
        else:
            # Revert back to the default limits
            if selected_option != "Behaviour Mean and SEM":
                self.ax.set_xlim(self.xlim_min, self.xlim_max)

            self.ax.set_ylim(self.default_y_limits)

        if not self.graph_settings_container_instance.x_axis_min_var.get() and not self.graph_settings_container_instance.y_axis_min_var.get():
            messagebox.showerror(
                "Error", "Please enter values for both x and y!")
            return

        self.figure_canvas.draw()

        self.settings_manager.save_variables()  # Save the values

        # Only close the popup if close=True
        if close and popup:
            popup.destroy()

    def fetch_behaviour_data(self):
        """
        Fetch the behaviour data from the table and return it as a list.

        Returns:
        - behaviour_occurrences: List of tuples containing start time, end time, pre-behaviour time, and post-behaviour time for each occurrence.
        - column_used: The column used for plotting.
        - pre_behaviour_times: List of pre-behaviour times.
        - post_behaviour_times: List of post-behaviour times.
        """
        graphed_behaviour = self.behaviour_choice_graph.get()

        if self.checkbox_state:
            column_used = 'baselined_z_score'
        else:
            column_used = self.selected_column_var.get()

        start_times = []
        end_times = []
        pre_behaviour_times = []
        post_behaviour_times = []
        behaviour_occurrences = []

        for index, row in self.tables[self.current_table_key].iterrows():
            current_behaviour_name = row['Behaviour Name']
            if current_behaviour_name == graphed_behaviour:
                behaviour_name = current_behaviour_name
                end_time_str = row['End Time']
                behaviour_type = row['Behaviour Type']
                start_time = round(float(row['Start Time']) / 60, 2)
                pre_behaviour_time = round(
                    float(row['Pre Behaviour Time']) / 60, 2)
                post_behaviour_time = round(
                    float(row['Post Behaviour Time']) / 60, 2)
                if behaviour_type == 'Point':
                    end_time = None
                else:
                    end_time = float(end_time_str) / 60
                start_times.append(start_time)
                end_times.append(end_time)
                pre_behaviour_times.append(pre_behaviour_time)
                post_behaviour_times.append(post_behaviour_time)
                behaviour_occurrences.append(
                    (start_time, end_time, pre_behaviour_time, post_behaviour_time))

        return behaviour_occurrences, column_used, pre_behaviour_times, post_behaviour_times

    def process_behaviour_data(self, behaviour_occurrences, column_used):
        """
        Process the behaviour data and return the processed data.

        Parameters:
        - behaviour_occurrences: List of tuples containing start time, end time, pre-behaviour time, and post-behaviour time for each occurrence.
        - column_used: The column used for plotting.

        Returns:
        - behaviour_data_by_instance: Dictionary with behaviour data for each instance.
        - time_points: Array of time points for the data.
        - start_time_adjusted: Adjusted start time.
        - end_time_adjusted: Adjusted end time.
        """
        behaviour_data_by_instance = {}
        time_points = []

        time_unit = self.graph_settings_container_instance.time_unit_menu.get()
        time_factor = self.get_time_scale(time_unit)

        reference_length = None
        for i, (start_time, end_time, pre_behaviour_time, post_behaviour_time) in enumerate(behaviour_occurrences):
            instance_id = f"Instance {i + 1}"
            start_point = start_time - pre_behaviour_time
            end_point = start_time + post_behaviour_time
            start_point_adjusted = start_point - 0.001
            if self.checkbox_state:
                selected_data_to_plot = self.dataframe.loc[
                    (self.dataframe['z_scored_time'] >= start_point_adjusted) &
                    (self.dataframe['z_scored_time'] <= end_point)
                ]
            else:
                selected_data_to_plot = self.dataframe.loc[
                    (self.dataframe.iloc[:, 0] >= start_point_adjusted) &
                    (self.dataframe.iloc[:, 0] <= end_point)
                ]
            behaviour_data = selected_data_to_plot[column_used].tolist()

            # Adjust to reference length
            if reference_length is None:
                reference_length = len(behaviour_data)
            else:
                behaviour_data = behaviour_data[:reference_length] if len(behaviour_data) > reference_length else behaviour_data + [np.nan] * (
                    reference_length - len(behaviour_data))

            time_points = np.linspace(
                start_point * time_factor, end_point * time_factor, len(behaviour_data))
            behaviour_data_by_instance[instance_id] = behaviour_data

            # Calculate the start time and end time
            start_time_adjusted = -pre_behaviour_time * time_factor
            end_time_adjusted = post_behaviour_time * time_factor

        return behaviour_data_by_instance, time_points, start_time_adjusted, end_time_adjusted

    def generate_behaviour_graph(self, ax, behaviour_data_by_instance, time_points, start_time_adjusted, end_time_adjusted):
        """
        Plot the mean and SEM of the behaviour data.

        Parameters:
        - ax: The axis on which to plot the data.
        - behaviour_data_by_instance: Dictionary with behaviour data for each instance.
        - time_points: Array of time points for the data.
        - start_time_adjusted: Adjusted start time.
        - end_time_adjusted: Adjusted end time.

        Returns:
        - mean_sem_df: DataFrame containing the mean and SEM of the behaviour data.
        """

        # Calculate mean and SEM
        mean_sem_df = self.generate_mean_sem_df(
            behaviour_data_by_instance, time_points, start_time_adjusted, end_time_adjusted)

        time_data = mean_sem_df['Time'].copy()
        converted_time_data, x_label = self.convert_and_retrieve_time(
            time_data, return_label=True)
        ax.set_xlabel(x_label)

        ax.plot(converted_time_data, mean_sem_df['Mean'], color=self.settings_manager.selected_trace_color, label='Mean',
                linewidth=float(self.graph_settings_container_instance.line_width_entry.get()))

        # Check if all SEM values are zero
        if not mean_sem_df['SEM'].eq(0).all():
            ax.fill_between(converted_time_data, mean_sem_df['Mean'] - mean_sem_df['SEM'], mean_sem_df['Mean'] + mean_sem_df['SEM'],
                            color=self.settings_manager.selected_sem_color,
                            alpha=0.5, label='SEM')

        # Only return mean_sem_df if it is needed in other parts of your code
        return mean_sem_df

    def add_duration_box(self, ax, mean_sem_df):
        """
        Add a duration box to the graph.

        Parameters:
        - ax: The axis on which to plot the duration box.
        - mean_sem_df: DataFrame containing the mean and SEM of the behaviour data.
        """
        if not self.graph_settings_container_instance.display_duration_box_var.get():
            return

        graphed_behaviour = self.behaviour_choice_graph.get()

        # Attempt to find the correct cache key
        # Simplified key, assuming it's just the behavior name
        cache_duration_key = graphed_behaviour

        if cache_duration_key in self.duration_data_cache:
            # Use cached values
            cached_data = self.duration_data_cache[cache_duration_key]
            mean_duration = cached_data["mean_duration"]
            sem_duration = cached_data["sem_duration"]
        else:
            # Handle the case where the data is not in the cache
            print(f"Duration data for {graphed_behaviour} not found in cache.")
            return

        # Convert the mean and SEM durations based on the time unit
        mean_duration = self.convert_and_retrieve_time(mean_duration)
        sem_duration = self.convert_and_retrieve_time(sem_duration)

        if self.graph_settings_container_instance.display_duration_box_var.get():
            height_modifier = float(
                self.graph_settings_container_instance.duration_box_placement.get())
            size_factor = float(
                self.graph_settings_container_instance.bar_graph_size_entry.get())

            # Calculate the y position of the duration box relative to the line graph
            max_mean = max(mean_sem_df['Mean'])
            min_mean = min(mean_sem_df['Mean'])
            mean_mean = mean_sem_df['Mean'].mean()
            if height_modifier >= 0:
                bar_y = mean_mean + (height_modifier * (max_mean - mean_mean))
            else:
                bar_y = mean_mean - (abs(height_modifier)
                                     * (mean_mean - min_mean))

            # Calculate the height of the bar relative to the range of mean
            mean_range = max_mean - min_mean
            bar_height = size_factor * mean_range

            # Draw the bar
            bars_container = ax.barh(bar_y, mean_duration, xerr=sem_duration, height=bar_height,
                                     color=self.graph_settings_container_instance.selected_bar_fill_color,
                                     edgecolor=self.graph_settings_container_instance.selected_bar_border_color,
                                     alpha=0.5,
                                     error_kw={'ecolor': self.graph_settings_container_instance.selected_bar_sem_color,
                                               'capsize': 5})

            # Store the bar items and error bars
            self.bar_items = list(bars_container.patches)
            self.bar_items.extend(bars_container.errorbar.get_children())

    def update_duration_box(self, ax=None, mean_sem_df=None):
        """
        Update the position of the duration box based on the Spinbox's value.

        Parameters:
        - ax: The axis on which to update the duration box (default is None).
        - mean_sem_df: DataFrame containing the mean and SEM of the behaviour data (default is None).

        Raises:
        - AttributeError: If the bar items do not have a 'remove' method.
        - ValueError: If the bar items are already gone.
        - KeyError: If the duration data is not found in the cache.
        """
        if not self.graph_settings_container_instance.display_duration_box_var.get():
            return

        ax = self.fig.gca()

        # Remove existing bar items
        if hasattr(self, 'bar_items'):
            for item in self.bar_items:
                try:
                    item.remove()
                except (AttributeError, ValueError):
                    pass  # Handle items that might not have a 'remove' method or are already gone

        # Identify the current behavior being processed
        current_behavior = self.behaviour_choice_graph.get()

        # Use the simplified cache key (just the behavior name)
        self.current_cache_key = current_behavior

        try:
            cached_data = self.duration_data_cache[self.current_cache_key]
            mean_duration = cached_data["mean_duration"]
            sem_duration = cached_data["sem_duration"]
            mean_sem_df = cached_data["mean_sem_df"]
        except KeyError:
            self.add_duration_box(ax, mean_sem_df)
            return

        # Convert the mean and SEM durations based on the time unit
        mean_duration = self.convert_and_retrieve_time(mean_duration)
        sem_duration = self.convert_and_retrieve_time(sem_duration)

        # Calculate position and size of bar
        max_mean = max(mean_sem_df['Mean'])
        min_mean = min(mean_sem_df['Mean'])
        mean_mean = mean_sem_df['Mean'].mean()

        height_modifier = float(
            self.graph_settings_container_instance.duration_box_placement.get())
        size_factor = float(
            self.graph_settings_container_instance.bar_graph_size_entry.get())

        if height_modifier >= 0:
            bar_y = mean_mean + (height_modifier * (max_mean - mean_mean))
        else:
            bar_y = mean_mean - (abs(height_modifier) * (mean_mean - min_mean))

        mean_range = max_mean - min_mean
        bar_height = size_factor * mean_range

        # Store the items on the axes before adding the bar
        old_items = set(ax.get_children())

        # Draw the bar
        bar = ax.barh(bar_y, mean_duration, xerr=sem_duration, height=bar_height, color=self.graph_settings_container_instance.selected_bar_fill_color,
                      edgecolor=self.graph_settings_container_instance.selected_bar_border_color, alpha=0.5,
                      error_kw={'ecolor': self.graph_settings_container_instance.selected_bar_sem_color, 'capsize': 5})

        # Store the items on the axes after adding the bar
        new_items = set(ax.get_children())

        # The new items (the bar and its error bars) are those that are in new_items but not in old_items
        self.bar_items = list(new_items - old_items)

        # Redraw the axes
        self.figure_canvas.draw_idle()

    def refresh_graph(self):
        """Redraw the graph based on current settings."""
        # Clear existing graph
        self.ax.clear()

        # Redraw the graph
        self.plot_full_trace(self.ax)

        # Refresh the canvas
        self.figure_canvas.draw()

    def handle_behaviour_change(self, *args, **kwargs):
        """Handle the behaviour change event."""
        selected_behaviour = self.graph_settings_container_instance.selected_behaviour_to_zero.get()
        # Check if the zero_x_axis_checkbox is checked
        if self.graph_settings_container_instance.zero_x_axis_checkbox_var.get() == 1 and (not selected_behaviour or selected_behaviour.strip() == ""):
            return

        # Redraw the plot or refresh the app display
        self.handle_figure_display_selection(None)

    def handle_zeroing(self, behaviours, start_times_min, end_times_min, converted_time_data):
        """
        Handle the zeroing of behaviour times and adjust the time data accordingly.

        Parameters:
        - behaviours: List of behaviours.
        - start_times_min: List of start times in minutes.
        - end_times_min: List of end times in minutes.
        - converted_time_data: Converted time data.

        Returns:
        - adjusted_start_times_min: List of adjusted start times in minutes.
        - adjusted_end_times_min: List of adjusted end times in minutes.
        - adjusted_time: Adjusted time data.
        """
        selected_behaviour = self.graph_settings_container_instance.selected_behaviour_to_zero.get()

        time_unit = self.graph_settings_container_instance.time_unit_menu.get()
        time_factor = self.get_time_scale(time_unit)

        # Make a copy of the converted_time_data to work with
        adjusted_time = converted_time_data.copy()

        # Adjust behavior times based on the selected behaviour
        adjusted_df = self.adjust_behavior_times(
            behaviours, start_times_min, end_times_min, selected_behaviour)

        if adjusted_df is None:
            print("Adjusted DataFrame is None. Returning original times.")
            return start_times_min, end_times_min, adjusted_time

        adjusted_start_times_min = adjusted_df['Adjusted Start Time'].tolist()
        adjusted_end_times_min = adjusted_df['Adjusted End Time'].tolist()

        # Calculate zero time based on the selected behaviour
        selected_behaviour_index = behaviours.index(selected_behaviour)
        zero_time = self.original_start_times_min[selected_behaviour_index] * time_factor

        # Adjust the converted time data based on the zeroing, ensuring correct units
        adjusted_time -= zero_time

        # Apply any additional offset based on the data being already adjusted or z-scoring
        if self.data_already_adjusted:
            if self.first_offset_time is None:
                self.first_offset_time = float(self.data_selection_frame.baseline_start_entry.get(
                )) / 60.0  # Convert seconds to minutes
            current_offset_time = float(self.data_selection_frame.baseline_start_entry.get(
            )) / 60.0  # Convert seconds to minutes
            offset_time = (current_offset_time -
                           self.first_offset_time) * time_factor
            adjusted_time += offset_time

        elif self.checkbox_state and self.figure_display_dropdown.get() == "Z-scored data":
            offset_time = (float(self.data_selection_frame.baseline_start_entry.get(
            )) / 60.0) * time_factor  # Convert seconds to minutes
            adjusted_time += offset_time

        return adjusted_start_times_min, adjusted_end_times_min, adjusted_time

    def adjust_behavior_times(self, behaviours, start_times_min, end_times_min, selected_behaviour):
        """
        Adjust behavior times based on the onset time of the selected behavior.

        Parameters:
        - behaviours: List of behaviours.
        - start_times_min: List of start times in minutes.
        - end_times_min: List of end times in minutes.
        - selected_behaviour: The selected behaviour to zero the times to.

        Returns:
        - adjusted_df: DataFrame with adjusted start and end times.
        """
        if selected_behaviour in behaviours and self.graph_settings_container_instance.zero_x_axis_checkbox_var.get() == 1:
            onset_index = behaviours.index(selected_behaviour)
            onset_time = start_times_min[onset_index]
        else:
            print(
                f"Selected behaviour {selected_behaviour} not found or checkbox not active.")
            return  # Return if behavior not found or checkbox is not active

        # Adjust the start and end times
        adjusted_start_times_min = [time_val -
                                    onset_time for time_val in start_times_min]
        adjusted_end_times_min = [
            time_val - onset_time if time_val is not None else None for time_val in end_times_min]

        # Create a DataFrame with the adjusted times
        adjusted_df = pd.DataFrame({
            'Behaviour Name': behaviours,
            'Adjusted Start Time': adjusted_start_times_min,
            'Adjusted End Time': adjusted_end_times_min
        })

        # Store the adjusted dataframe in the dictionary
        self.adjusted_behaviour_dataframes[selected_behaviour] = adjusted_df

        return adjusted_df

    def save_image(self):
        """Saves the current figure to a file, applying user-defined font and label settings."""
        # Get current labels from the figure
        current_xlabel = self.fig.axes[0].get_xlabel()
        current_ylabel = self.fig.axes[0].get_ylabel()

        # Apply user-defined font sizes, preserving existing labels if no new label is specified
        xlabel_fontsize = self.export_options_container.font_settings.get(
            'xlabel_fontsize')
        ylabel_fontsize = self.export_options_container.font_settings.get(
            'ylabel_fontsize')
        xtick_fontsize = self.export_options_container.font_settings.get(
            'xtick_fontsize')
        ytick_fontsize = self.export_options_container.font_settings.get(
            'ytick_fontsize')
        # Assuming self.graph_settings exists and is properly initialized
        y_axis_name = self.export_options_container.font_settings.get(
            'y_axis_name', '')
        y_label_to_use = y_axis_name if y_axis_name else current_ylabel

        # Retrieve height and width from entries
        height_str = self.export_options_container.height_entry.get().strip()
        width_str = self.export_options_container.width_entry.get().strip()
        fig_copy = copy.deepcopy(self.fig)

        # Only set figure size if both height and width are provided
        if height_str and width_str:
            try:
                height = float(height_str)
                width = float(width_str)
                # Set figure size based on the user input
                fig_copy.set_size_inches(width, height)
            except ValueError:
                # If conversion fails, possibly log or alert the user, but proceed to save with default size
                print("Invalid height or width. Using default figure size.")

        fig_width, fig_height = fig_copy.get_size_inches()

        # Calculate the diagonal length of the figure
        diagonal_length = math.sqrt(fig_width**2 + fig_height**2)

        # Set the default font size relative to the diagonal length of the figure
        default_font_size = 1.2 * diagonal_length

        xlabel_fontsize = self.export_options_container.font_settings.get(
            'xlabel_fontsize')
        ylabel_fontsize = self.export_options_container.font_settings.get(
            'ylabel_fontsize')
        xtick_fontsize = self.export_options_container.font_settings.get(
            'xtick_fontsize')
        ytick_fontsize = self.export_options_container.font_settings.get(
            'ytick_fontsize')
        # Assuming self.graph_settings exists and is properly initialized
        y_axis_name = self.export_options_container.font_settings.get(
            'y_axis_name', '')

        # Apply the calculated font size
        fig_copy.axes[0].set_xlabel(current_xlabel, fontsize=int(
            xlabel_fontsize) if xlabel_fontsize else default_font_size)
        fig_copy.axes[0].set_ylabel(y_label_to_use, fontsize=int(
            ylabel_fontsize) if ylabel_fontsize else default_font_size)
        fig_copy.axes[0].xaxis.set_tick_params(labelsize=int(
            xtick_fontsize) if xtick_fontsize else default_font_size)
        fig_copy.axes[0].yaxis.set_tick_params(labelsize=int(
            ytick_fontsize) if ytick_fontsize else default_font_size)

        fig_copy.tight_layout()

        selected_format = self.export_options_container.image_format_combobox.get().lower()
        # Get the DPI value from the entry box
        dpi = int(self.export_options_container.dpi_entry.get())
        file_path = self.file_path_var.get()  # Get the file path
        figure_display = self.figure_display_dropdown.get()

        # If figure display type is "Behaviour Mean and SEM", add the behaviour choice to the base name
        if figure_display == "Behaviour Mean and SEM":
            behaviour_choice = self.behaviour_choice_graph.get()
            base_name = f"{self.mouse_name}_{
                figure_display}_{behaviour_choice}"
        else:
            base_name = f"{self.mouse_name}_{figure_display}"

        if self.mouse_name is None or self.mouse_name == "":
            base_name, _ = os.path.splitext(os.path.basename(
                file_path))  # Extract base name from file path

        # Create "exported images" folder within the directory the data file came from
        dir_name = os.path.dirname(file_path)
        exported_images_dir = os.path.join(
            dir_name, f"exported_images_{self.mouse_name}")
        # Creates the directory if it doesn't exist
        os.makedirs(exported_images_dir, exist_ok=True)

        # Add a simpler timestamp to the base name (just date)
        timestamp = datetime.datetime.now().strftime("%b%d_%H%M")  # Oct23_1530 format
        base_name = f"{base_name}_{timestamp}"

        # Check if file already exists. If it does, add a suffix
        counter = 1
        original_base_name = base_name

        while os.path.isfile(os.path.join(exported_images_dir, f"{base_name}.{selected_format}")):
            base_name = f"{original_base_name}_{counter}"
            counter += 1

        filename = os.path.join(exported_images_dir,
                                f"{base_name}.{selected_format}")
        fig_copy.savefig(filename, format=selected_format, dpi=dpi)

        plt.close(fig_copy)
